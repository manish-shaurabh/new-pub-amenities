"""
Reports Builder — Superadmin-only customisable report engine.

A single generic engine that handles:
  metrics    : pct_working | mttr | defect_frequency | rejection_rate | inspection_volume
  dimensions : station | location | department | asset_type | supervisor | time_bucket
  cross-tab  : up to 2 dimensions (dim_x × dim_y → matrix cells)
  filters    : station_ids, dept_ids, asset_type_ids, role, list_type, status
  windows    : last_7d | last_30d | last_90d | last_180d | fy | all_time | custom

Endpoints
  POST /api/reports/builder/run/{user_id}      — execute a config, return rows/matrix
  GET  /api/reports/builder/saved/{user_id}    — list saved configs
  POST /api/reports/builder/save/{user_id}     — save a config
  DELETE /api/reports/builder/saved/{id}/{user_id}
  GET  /api/reports/builder/featured           — 8 ready-made configs
  GET  /api/reports/builder/dimensions/{user_id} — filter option metadata
  POST /api/reports/builder/export/{fmt}/{user_id} (fmt=pdf|excel|csv)

All endpoints reject non-superadmin users.
"""
import io
import csv
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from bson import ObjectId
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from database import (
    users_collection, assets_collection, asset_types_collection,
    stations_collection, locations_collection, departments_collection,
    orange_list_collection, inspections_collection, saved_reports_collection,
    report_runs_collection, saved_dossiers_collection, now_ist,
)

router = APIRouter()


# ════════════════════════════════════════════════════════════════════════════
# Auth gate — Superadmin only
# ════════════════════════════════════════════════════════════════════════════
async def _ensure_sa(user_id: str) -> dict:
    try:
        u = await users_collection.find_one({"_id": ObjectId(user_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user_id")
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    if u.get("role") != "superadmin":
        raise HTTPException(status_code=403, detail="Superadmin only")
    return u


# ════════════════════════════════════════════════════════════════════════════
# Datetime helper (handles naive IST strings/datetimes uniformly)
# ════════════════════════════════════════════════════════════════════════════
def _parse_dt(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v.replace("Z", "").replace("+00:00", ""))
        except Exception:
            return None
    return None


# ════════════════════════════════════════════════════════════════════════════
# Time window
# ════════════════════════════════════════════════════════════════════════════
def _window(window: str, custom_from: Optional[str], custom_to: Optional[str]):
    """Return (from_dt, to_dt) — both naive IST, inclusive."""
    now = now_ist()
    if window == "all_time":
        return None, now
    if window == "custom":
        return _parse_dt(custom_from), _parse_dt(custom_to) or now
    if window == "fy":
        # Indian FY: Apr 1 – Mar 31
        if now.month >= 4:
            start = datetime(now.year, 4, 1)
        else:
            start = datetime(now.year - 1, 4, 1)
        return start, now
    days = {"last_7d": 7, "last_30d": 30, "last_90d": 90, "last_180d": 180}.get(window)
    if days:
        return now - timedelta(days=days), now
    return None, now


# ════════════════════════════════════════════════════════════════════════════
# Universe loader (single trip — small dataset)
# ════════════════════════════════════════════════════════════════════════════
async def _load_universe(filters: dict):
    """Pre-filter assets/users/etc. Returns a dict of dicts indexed by _id."""
    assets = await assets_collection.find({}).to_list(20000)
    types = await asset_types_collection.find({}).to_list(2000)
    stations = await stations_collection.find({}).to_list(2000)
    locations = await locations_collection.find({}).to_list(5000)
    depts = await departments_collection.find({}).to_list(500)
    users = await users_collection.find({}).to_list(5000)
    type_by_id = {str(t["_id"]): t for t in types}
    station_by_id = {str(s["_id"]): s for s in stations}
    location_by_id = {str(l["_id"]): l for l in locations}
    dept_by_id = {str(d["_id"]): d for d in depts}
    user_by_id = {str(u["_id"]): u for u in users}

    # Station → list of RO/SUP user ids (multiple ROs/SUPs can map to one station)
    ros_by_station: Dict[str, List[str]] = defaultdict(list)
    sups_by_station: Dict[str, List[str]] = defaultdict(list)
    for u in users:
        sids = u.get("assigned_stations") or []
        if u.get("role") == "reporting_officer":
            for sid in sids:
                ros_by_station[sid].append(str(u["_id"]))
        elif u.get("role") == "supervisor":
            for sid in sids:
                sups_by_station[sid].append(str(u["_id"]))

    # All open OL entries — used to determine list_type for currently-defective assets
    open_ols = await orange_list_collection.find(
        {"status": {"$ne": "resolved"}}).to_list(20000)
    open_ol_by_asset = {ol["asset_id"]: ol for ol in open_ols}

    # Apply filters at asset level
    f_stations = set(filters.get("station_ids") or [])
    f_depts = set(filters.get("dept_ids") or [])
    f_types = set(filters.get("asset_type_ids") or [])
    f_statuses = set(filters.get("asset_statuses") or [])
    if f_stations or f_depts or f_types or f_statuses:
        def _ok(a):
            if f_stations and a.get("station_id") not in f_stations:
                return False
            t = type_by_id.get(a.get("asset_type_id"), {})
            if f_depts and t.get("department_id") not in f_depts:
                return False
            if f_types and a.get("asset_type_id") not in f_types:
                return False
            if f_statuses and a.get("status") not in f_statuses:
                return False
            return True
        assets = [a for a in assets if _ok(a)]

    return {
        "assets": assets,
        "type_by_id": type_by_id,
        "station_by_id": station_by_id,
        "location_by_id": location_by_id,
        "dept_by_id": dept_by_id,
        "user_by_id": user_by_id,
        "ros_by_station": dict(ros_by_station),
        "sups_by_station": dict(sups_by_station),
        "open_ol_by_asset": open_ol_by_asset,
    }


# ════════════════════════════════════════════════════════════════════════════
# Dimension extractors — return (key, label) for an asset/event row
# ════════════════════════════════════════════════════════════════════════════
def _dim_for_asset(U: dict, asset: dict, dim: str):
    if dim == "station":
        s = U["station_by_id"].get(asset.get("station_id"), {})
        return asset.get("station_id") or "—", s.get("name", "—")
    if dim == "location":
        l = U["location_by_id"].get(asset.get("location_id"), {})
        return asset.get("location_id") or "—", l.get("name", "—")
    if dim == "department":
        t = U["type_by_id"].get(asset.get("asset_type_id"), {})
        d = U["dept_by_id"].get(t.get("department_id"), {})
        return t.get("department_id") or "—", d.get("name", "—")
    if dim == "asset_type":
        t = U["type_by_id"].get(asset.get("asset_type_id"), {})
        return asset.get("asset_type_id") or "—", t.get("name", "—")
    if dim == "asset":
        return str(asset.get("_id") or "—"), asset.get("asset_number") or "—"
    if dim == "ro":
        ro_ids = U.get("ros_by_station", {}).get(asset.get("station_id"), [])
        if not ro_ids:
            return "—", "— (unassigned)"
        u = U["user_by_id"].get(ro_ids[0], {})
        return ro_ids[0], u.get("name") or u.get("employee_id") or "—"
    if dim == "asup":
        s = U["station_by_id"].get(asset.get("station_id"), {})
        asup_id = s.get("approving_supervisor_id")
        if not asup_id:
            return "—", "— (unassigned)"
        u = U["user_by_id"].get(asup_id, {})
        return asup_id, u.get("name") or u.get("employee_id") or "—"
    if dim == "list_type":
        ol = U.get("open_ol_by_asset", {}).get(str(asset["_id"]))
        if not ol:
            # Fall back to status-derived label
            st = asset.get("status")
            if st == "working":
                return "working", "Working"
            if st == "pending_approval":
                return "yellow", "Yellow (pending)"
            return "—", "— (no open OL)"
        lt = (ol.get("list_type") or "").lower() or "orange"
        return lt, lt.capitalize()
    return "—", "—"


def _dim_for_user(U: dict, user_id: str, dim: str):
    """For event-level dims that resolve to a user (inspector/reporter/etc)."""
    if not user_id:
        return ("—", "— (unattributed)")
    u = U["user_by_id"].get(user_id, {})
    return user_id, u.get("name") or u.get("employee_id") or "—"


def _dim_for_time(dt: datetime, dim: str):
    if not dt:
        return None
    if dim == "time_bucket_day":
        return dt.strftime("%Y-%m-%d"), dt.strftime("%d %b %Y")
    if dim == "time_bucket_week":
        y, w, _ = dt.isocalendar()
        return f"{y}-W{w:02d}", f"{y} W{w:02d}"
    if dim == "time_bucket_month":
        return dt.strftime("%Y-%m"), dt.strftime("%b %Y")
    if dim == "time_bucket_quarter":
        q = (dt.month - 1) // 3 + 1
        return f"{dt.year}-Q{q}", f"{dt.year} Q{q}"
    return None


_AGE_BANDS = [(0, 2, "0-2h"), (2, 8, "2-8h"), (8, 24, "8-24h"),
              (24, 72, "1-3d"), (72, 168, "3-7d"), (168, 1e9, "7d+")]


def _band_for_hours(hrs: float):
    for lo, hi, lbl in _AGE_BANDS:
        if lo <= hrs < hi:
            return lbl
    return "—"


def _dim_for_event(event: dict, dim: str):
    """Event-level dimensions: age bands, hour-of-day, day-of-week."""
    if dim == "defect_age_band":
        # For events whose value is the defect-at-event age (only applies when event carries _defect_age_hrs)
        hrs = event.get("_defect_age_hrs")
        if hrs is None:
            return None
        lbl = _band_for_hours(hrs)
        return lbl, lbl
    if dim == "repair_age_band":
        hrs = event.get("_value")  # for MTTR events, value=hours
        if hrs is None:
            return None
        lbl = _band_for_hours(hrs)
        return lbl, lbl
    if dim == "hour_of_day":
        dt = event.get("_event_dt")
        if not dt:
            return None
        h = dt.hour
        return f"{h:02d}", f"{h:02d}:00"
    if dim == "day_of_week":
        dt = event.get("_event_dt")
        if not dt:
            return None
        names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        return str(dt.weekday()), names[dt.weekday()]
    if dim == "list_type":
        # Event-level list_type if carried explicitly on the OL event
        lt = (event.get("_list_type") or "").lower()
        if not lt:
            return None
        return lt, lt.capitalize()
    return None


def _key_label(U: dict, asset: Optional[dict], event: Optional[dict],
               dim: str) -> tuple:
    """Resolve a (key, label) tuple for any dimension+row combination."""
    if dim.startswith("time_bucket"):
        dt = event.get("_event_dt") if event else None
        return _dim_for_time(dt, dim) or ("—", "—")
    if dim in ("defect_age_band", "repair_age_band", "hour_of_day", "day_of_week"):
        return _dim_for_event(event or {}, dim) or ("—", "—")
    if dim == "supervisor":
        actor = (event or {}).get("_actor_id")
        return _dim_for_user(U, actor, dim)
    if dim == "inspector":
        actor = (event or {}).get("_inspector_id") or (event or {}).get("_actor_id")
        return _dim_for_user(U, actor, dim)
    if dim == "reporter":
        actor = (event or {}).get("_reporter_id")
        return _dim_for_user(U, actor, dim)
    # list_type can be event-level if available, else fall back to asset-level
    if dim == "list_type" and event and event.get("_list_type"):
        r = _dim_for_event(event, dim)
        if r:
            return r
    if asset:
        return _dim_for_asset(U, asset, dim)
    return ("—", "—")


# ════════════════════════════════════════════════════════════════════════════
# Metric engine — each metric returns a list of "events", each event is a row
# that contributes to one cell. Then we aggregate per (key_x, key_y).
# ════════════════════════════════════════════════════════════════════════════
def _classify(asset, ol_open):
    st = asset.get("status")
    if st == "working":
        return "working"
    if st == "pending_approval":
        return "yellow"
    ds = (ol_open or {}).get("defective_since") or asset.get("defective_since")
    ds_dt = _parse_dt(ds)
    if not ds_dt:
        return "orange"
    hours = (now_ist() - ds_dt).total_seconds() / 3600
    return "red" if hours > 24 else "orange"


def _percentile(arr: List[float], p: float):
    if not arr:
        return None
    arr = sorted(arr)
    k = (len(arr) - 1) * p
    f = int(k); c = min(f + 1, len(arr) - 1)
    if f == c:
        return arr[f]
    return arr[f] + (arr[c] - arr[f]) * (k - f)


async def _events_pct_working(U: dict, win, filters):
    """Each asset in scope is one event with class=working/yellow/orange/red."""
    open_ols = await orange_list_collection.find(
        {"status": {"$ne": "resolved"}}).to_list(20000)
    ol_by_asset = {ol["asset_id"]: ol for ol in open_ols}
    events = []
    for a in U["assets"]:
        cls = _classify(a, ol_by_asset.get(str(a["_id"])))
        events.append({"_asset": a, "_value": 1 if cls in ("working", "yellow") else 0,
                       "_class": cls})
    return events


async def _events_mttr(U: dict, win, filters):
    """Resolved OL entries inside the window. value = repair hours."""
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    repair_cap = filters.get("repair_cap_hours")  # exclude > N hours if set
    include_rejected = bool(filters.get("include_rejected_in_mttr"))
    ols = await orange_list_collection.find({"status": "resolved"}).to_list(50000)
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    events = []
    for ol in ols:
        if ol.get("asset_id") not in asset_ids:
            continue
        ds = _parse_dt(ol.get("defective_since"))
        mw = _parse_dt(ol.get("marked_working_at"))
        if not ds or not mw:
            continue
        if f_dt and mw < f_dt:
            continue
        if t_dt and mw > t_dt:
            continue
        hours = (mw - ds).total_seconds() / 3600
        if hours < 0:
            continue
        if repair_cap and hours > float(repair_cap):
            continue
        if not include_rejected and ol.get("rejected_at"):
            # Skip OLs that had any rejected mark-working in their history
            continue
        events.append({
            "_asset": asset_by_id.get(ol.get("asset_id")),
            "_value": hours,
            "_event_dt": mw,
            "_actor_id": ol.get("marked_working_by"),
            "_reporter_id": ol.get("reported_by"),
            "_list_type": ol.get("list_type") or "",
            "_resolved_without_rejection": not bool(ol.get("rejected_at")),
        })
    return events


async def _events_defect_frequency(U: dict, win, filters):
    """Each new OL entry inside window = 1 event."""
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    ols = await orange_list_collection.find({}).to_list(50000)
    events = []
    for ol in ols:
        if ol.get("asset_id") not in asset_ids:
            continue
        ds = _parse_dt(ol.get("defective_since")) or _parse_dt(ol.get("created_at"))
        if not ds:
            continue
        if f_dt and ds < f_dt:
            continue
        if t_dt and ds > t_dt:
            continue
        events.append({
            "_asset": asset_by_id.get(ol.get("asset_id")),
            "_value": 1,
            "_event_dt": ds,
            "_actor_id": ol.get("reported_by"),
        })
    return events


async def _events_rejection_rate(U: dict, win, filters):
    """Each closed (approved or rejected) verdict in window = 1 event.
    _value = 1 if rejected else 0."""
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    ols = await orange_list_collection.find({"status": {"$in": ["resolved", "defective"]}}).to_list(50000)
    events = []
    for ol in ols:
        if ol.get("asset_id") not in asset_ids:
            continue
        verdict = None
        verdict_dt = None
        actor = None
        if ol.get("status") == "resolved" and ol.get("approved_at"):
            verdict = 0
            verdict_dt = _parse_dt(ol.get("approved_at"))
            actor = ol.get("approved_by")
        elif ol.get("rejection_remarks") or ol.get("rejected_at"):
            verdict = 1
            verdict_dt = _parse_dt(ol.get("rejected_at"))
            actor = ol.get("rejected_by") or ol.get("approved_by")
        if verdict is None or not verdict_dt:
            continue
        if f_dt and verdict_dt < f_dt:
            continue
        if t_dt and verdict_dt > t_dt:
            continue
        events.append({
            "_asset": asset_by_id.get(ol.get("asset_id")),
            "_value": verdict,
            "_event_dt": verdict_dt,
            "_actor_id": actor,
        })
    return events


async def _events_inspection_volume(U: dict, win, filters):
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    inspections = await inspections_collection.find({}).to_list(50000)
    events = []
    for ins in inspections:
        ins_dt = _parse_dt(ins.get("inspection_at"))
        if not ins_dt:
            continue
        if f_dt and ins_dt < f_dt:
            continue
        if t_dt and ins_dt > t_dt:
            continue
        for it in ins.get("items", []):
            if it.get("asset_id") not in asset_ids:
                continue
            events.append({
                "_asset": asset_by_id.get(it.get("asset_id")),
                "_value": 1,
                "_event_dt": ins_dt,
                "_actor_id": ins.get("inspector_id"),
            })
    return events


METRIC_FNS = {
    "pct_working": _events_pct_working,
    "mttr": _events_mttr,
    "defect_frequency": _events_defect_frequency,
    "rejection_rate": _events_rejection_rate,
    "inspection_volume": _events_inspection_volume,
}


# ──── Layer 2 metrics ─────────────────────────────────────────────────────
async def _events_first_time_fix(U, win, filters):
    """For each resolved OL in window, value = 1 if no rejected mark-working ever, else 0."""
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    ols = await orange_list_collection.find({"status": "resolved"}).to_list(50000)
    events = []
    for ol in ols:
        if ol.get("asset_id") not in asset_ids:
            continue
        mw = _parse_dt(ol.get("marked_working_at"))
        if not mw:
            continue
        if f_dt and mw < f_dt: continue
        if t_dt and mw > t_dt: continue
        events.append({
            "_asset": asset_by_id.get(ol.get("asset_id")),
            "_value": 0 if ol.get("rejected_at") else 1,
            "_event_dt": mw,
            "_actor_id": ol.get("marked_working_by"),
        })
    return events


async def _events_recurrence(U, win, filters):
    """% of assets in scope that defected, were resolved, and re-defected within
    `recurrence_within_days` (default 30) days inside the window.
    Each asset with at least one resolved-then-redefected pair = 1, else 0."""
    f_dt, t_dt = win
    days = float(filters.get("recurrence_within_days") or 30)
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    ols = await orange_list_collection.find({}).to_list(50000)
    by_asset: Dict[str, list] = defaultdict(list)
    for ol in ols:
        aid = ol.get("asset_id")
        if aid in asset_ids:
            by_asset[aid].append(ol)
    events = []
    for aid, aol in by_asset.items():
        # Sort by defective_since
        aol.sort(key=lambda x: _parse_dt(x.get("defective_since")) or datetime.min)
        recur = 0
        for i, ol in enumerate(aol[:-1]):
            mw = _parse_dt(ol.get("marked_working_at"))
            if not mw:
                continue
            nxt_ds = _parse_dt(aol[i+1].get("defective_since"))
            if not nxt_ds:
                continue
            if (nxt_ds - mw).total_seconds() / 86400 > days:
                continue
            # Use nxt_ds as the event time
            if f_dt and nxt_ds < f_dt: continue
            if t_dt and nxt_ds > t_dt: continue
            recur = 1
            events.append({
                "_asset": asset_by_id.get(aid),
                "_value": 1,
                "_event_dt": nxt_ds,
            })
            break
        if recur == 0 and asset_by_id.get(aid):
            # Anchor a "0" event so denominator includes this asset
            events.append({
                "_asset": asset_by_id.get(aid),
                "_value": 0,
                "_event_dt": now_ist(),
            })
    return events


async def _events_backlog_age(U, win, filters):
    """For each currently-OPEN OL entry, value = hours since defective_since."""
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    open_ols = await orange_list_collection.find({"status": {"$ne": "resolved"}}).to_list(50000)
    now = now_ist()
    events = []
    for ol in open_ols:
        if ol.get("asset_id") not in asset_ids:
            continue
        ds = _parse_dt(ol.get("defective_since"))
        if not ds:
            continue
        hrs = (now - ds).total_seconds() / 3600
        if hrs < 0:
            continue
        events.append({
            "_asset": asset_by_id.get(ol.get("asset_id")),
            "_value": hrs,
            "_event_dt": ds,
            "_list_type": ol.get("list_type") or "",
            "_defect_age_hrs": hrs,
        })
    return events


async def _events_throughput(U, win, filters):
    """Count of resolved OLs per period (same shape as defect_frequency but for resolutions)."""
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    ols = await orange_list_collection.find({"status": "resolved"}).to_list(50000)
    events = []
    for ol in ols:
        if ol.get("asset_id") not in asset_ids: continue
        mw = _parse_dt(ol.get("marked_working_at"))
        if not mw: continue
        if f_dt and mw < f_dt: continue
        if t_dt and mw > t_dt: continue
        events.append({
            "_asset": asset_by_id.get(ol.get("asset_id")),
            "_value": 1,
            "_event_dt": mw,
            "_actor_id": ol.get("marked_working_by"),
        })
    return events


async def _events_avg_approval_lag(U, win, filters):
    """Hours between marked_working_at and approved_at."""
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    asset_by_id = {str(a["_id"]): a for a in U["assets"]}
    ols = await orange_list_collection.find({"status": "resolved",
                                             "approved_at": {"$exists": True}}).to_list(50000)
    events = []
    for ol in ols:
        if ol.get("asset_id") not in asset_ids: continue
        mw = _parse_dt(ol.get("marked_working_at"))
        ap = _parse_dt(ol.get("approved_at"))
        if not mw or not ap: continue
        if f_dt and ap < f_dt: continue
        if t_dt and ap > t_dt: continue
        hrs = (ap - mw).total_seconds() / 3600
        if hrs < 0: continue
        events.append({
            "_asset": asset_by_id.get(ol.get("asset_id")),
            "_value": hrs,
            "_event_dt": ap,
            "_actor_id": ol.get("approved_by"),
        })
    return events


async def _events_pct_pending(U, win, filters):
    """For every asset, value = 1 if currently pending_approval (yellow) else 0."""
    events = []
    for a in U["assets"]:
        events.append({
            "_asset": a,
            "_value": 1 if a.get("status") == "pending_approval" else 0,
        })
    return events


async def _events_inspection_coverage(U, win, filters):
    """For each asset, value = 1 if inspected at least once in window, else 0."""
    f_dt, t_dt = win
    asset_ids = {str(a["_id"]) for a in U["assets"]}
    inspections = await inspections_collection.find({}).to_list(50000)
    inspected_in_win = set()
    for ins in inspections:
        ins_dt = _parse_dt(ins.get("inspection_at"))
        if not ins_dt: continue
        if f_dt and ins_dt < f_dt: continue
        if t_dt and ins_dt > t_dt: continue
        for it in ins.get("items", []):
            if it.get("asset_id") in asset_ids:
                inspected_in_win.add(it.get("asset_id"))
    events = []
    for a in U["assets"]:
        events.append({
            "_asset": a,
            "_value": 1 if str(a["_id"]) in inspected_in_win else 0,
        })
    return events


METRIC_FNS.update({
    "first_time_fix_rate": _events_first_time_fix,
    "recurrence_rate": _events_recurrence,
    "backlog_age": _events_backlog_age,
    "throughput": _events_throughput,
    "avg_approval_lag": _events_avg_approval_lag,
    "pct_pending": _events_pct_pending,
    "inspection_coverage": _events_inspection_coverage,
})


# ════════════════════════════════════════════════════════════════════════════
# Aggregator — combine events into cells per (dim_x, dim_y)
# ════════════════════════════════════════════════════════════════════════════
def _aggregate(events: List[dict], metric: str, U: dict,
               dim_x: str, dim_y: Optional[str],
               output: Optional[dict] = None) -> dict:
    """Combine events into rows / matrix, then apply Layer 4 output controls."""
    output = output or {}
    cells: Dict[tuple, List[float]] = defaultdict(list)
    label_x: Dict[str, str] = {}
    label_y: Dict[str, str] = {}

    for e in events:
        kx, lx = _key_label(U, e.get("_asset"), e, dim_x)
        ky = ly = None
        if dim_y:
            ky, ly = _key_label(U, e.get("_asset"), e, dim_y)
        if not kx:
            continue
        label_x[kx] = lx
        if dim_y and ky:
            label_y[ky] = ly
            cells[(kx, ky)].append(e["_value"])
        else:
            cells[(kx, None)].append(e["_value"])

    # Pct-style metrics (rate as 0/1 fraction)
    PCT_METRICS = {"pct_working", "rejection_rate", "first_time_fix_rate",
                   "recurrence_rate", "pct_pending", "inspection_coverage"}
    SUM_METRICS = {"defect_frequency", "inspection_volume", "throughput"}
    HRS_METRICS = {"mttr", "backlog_age", "avg_approval_lag"}

    def _stat(arr):
        n = len(arr)
        if n == 0:
            return {"n": 0, "value": None}
        if metric in PCT_METRICS:
            return {"n": n, "value": round(sum(arr) / n * 100, 1)}
        if metric in HRS_METRICS:
            arr_s = sorted(arr)
            return {
                "n": n,
                "value": round(_percentile(arr_s, 0.5), 1),  # median
                "p25": round(_percentile(arr_s, 0.25), 1),
                "p75": round(_percentile(arr_s, 0.75), 1),
                "p90": round(_percentile(arr_s, 0.90), 1),
                "p99": round(_percentile(arr_s, 0.99), 1),
                "min": round(min(arr_s), 1),
                "max": round(max(arr_s), 1),
                "mean": round(sum(arr) / n, 1),
            }
        if metric in SUM_METRICS:
            return {"n": n, "value": int(sum(arr))}
        return {"n": n, "value": None}

    # ──── single-dimension path ────
    if not dim_y:
        rows = []
        for kx, vals in cells.items():
            stat = _stat(vals)
            rows.append({"key_x": kx[0], "label_x": label_x.get(kx[0], "—"), **stat})
        rows = _apply_output_controls(rows, metric, output, dim_x)
        return {"rows": rows}

    # ──── 2-dim matrix path ────
    row_keys = sorted(label_x.keys(), key=lambda k: label_x.get(k, ""))
    col_keys = sorted(label_y.keys(), key=lambda k: label_y.get(k, ""))
    matrix = []
    for rk in row_keys:
        row_cells = []
        for ck in col_keys:
            vals = cells.get((rk, ck), [])
            row_cells.append(_stat(vals))
        matrix.append(row_cells)

    # n-threshold greys out cells (mark as None with `_n_below: True`)
    n_thr = int(output.get("n_threshold") or 0)
    if n_thr > 0:
        for r in matrix:
            for c in r:
                if c.get("n", 0) < n_thr:
                    c["_n_below_threshold"] = True
                    c["value"] = None

    return {
        "row_keys": row_keys,
        "row_labels": [label_x.get(k, "—") for k in row_keys],
        "col_keys": col_keys,
        "col_labels": [label_y.get(k, "—") for k in col_keys],
        "matrix": matrix,
    }


# ──── Layer 4 — output controls (sort, top_n, totals row, n threshold, Other) ────
def _apply_output_controls(rows: list, metric: str, output: dict, dim_x: str) -> list:
    n_thr = int(output.get("n_threshold") or 0)
    if n_thr > 0:
        rows = [r for r in rows if r.get("n", 0) >= n_thr]

    # Sort
    sort_by = (output.get("sort_by") or "value").lower()
    sort_dir = (output.get("sort_dir") or "desc").lower()
    rev = (sort_dir == "desc")
    if sort_by == "label":
        rows.sort(key=lambda r: r.get("label_x") or "", reverse=rev)
    elif sort_by == "n":
        rows.sort(key=lambda r: r.get("n") or 0, reverse=rev)
    else:
        rows.sort(key=lambda r: (r.get("value") or 0), reverse=rev)

    # Top N + bucket-Other
    top_n = output.get("top_n")
    bucket_after = output.get("bucket_other_after")
    if bucket_after and len(rows) > int(bucket_after):
        head, tail = rows[:int(bucket_after)], rows[int(bucket_after):]
        if tail:
            tail_n = sum(r.get("n", 0) for r in tail)
            tail_vals = []
            for r in tail:
                if r.get("value") is not None and r.get("n", 0) > 0:
                    tail_vals.extend([r["value"]] * r["n"])
            agg_value = round(sum(tail_vals) / len(tail_vals), 1) if tail_vals else None
            head.append({"key_x": "__other__", "label_x": f"Other ({len(tail)})",
                         "n": tail_n, "value": agg_value, "_is_other": True})
        rows = head
    if top_n:
        rows = rows[:int(top_n)]

    # Totals row
    if output.get("totals_row"):
        all_n = sum(r.get("n", 0) for r in rows if r.get("key_x") != "__other__")
        if metric in {"pct_working", "rejection_rate", "first_time_fix_rate",
                      "recurrence_rate", "pct_pending", "inspection_coverage"}:
            # Weighted average pct
            num = sum((r.get("value") or 0) * r.get("n", 0) for r in rows)
            total_value = round(num / all_n, 1) if all_n else None
        elif metric in {"defect_frequency", "inspection_volume", "throughput"}:
            total_value = sum((r.get("value") or 0) for r in rows)
        else:
            # MTTR-style: report mean of medians (informational)
            vals = [r.get("value") for r in rows if r.get("value") is not None]
            total_value = round(sum(vals) / len(vals), 1) if vals else None
        rows.append({"key_x": "__total__", "label_x": "Total (weighted)",
                     "n": all_n, "value": total_value, "_is_total": True})
    return rows


# ════════════════════════════════════════════════════════════════════════════
# Run engine
# ════════════════════════════════════════════════════════════════════════════
class BuilderConfig(BaseModel):
    metric: str
    dim_x: str
    dim_y: Optional[str] = None
    window: str = "last_30d"
    custom_from: Optional[str] = None
    custom_to: Optional[str] = None
    # Filters: station_ids, dept_ids, asset_type_ids, asset_statuses,
    #          inspection_statuses, repair_cap_hours, recurrence_within_days,
    #          include_rejected_in_mttr, hour_from, hour_to, list_types
    filters: Dict[str, Any] = Field(default_factory=dict)
    # Layer 4 output controls
    output: Dict[str, Any] = Field(default_factory=dict)
    # Layer 4 visualisation hint (frontend uses this; backend ignores)
    viz: Optional[str] = None
    # Layer 4 annotations
    annotations: Optional[Dict[str, Any]] = None
    # Compare-to flag (Layer 4) — when True, run engine for previous period and emit deltas
    compare_to_previous: bool = False


async def _run_config(cfg: BuilderConfig) -> dict:
    if cfg.metric not in METRIC_FNS:
        raise HTTPException(status_code=400, detail=f"Unknown metric '{cfg.metric}'")
    win = _window(cfg.window, cfg.custom_from, cfg.custom_to)
    U = await _load_universe(cfg.filters or {})
    events = await METRIC_FNS[cfg.metric](U, win, cfg.filters or {})
    # Layer 3: hour-of-day filter (post-event filtering)
    hf = (cfg.filters or {}).get("hour_from")
    ht = (cfg.filters or {}).get("hour_to")
    if hf is not None and ht is not None:
        hf, ht = int(hf), int(ht)
        def _in(h):
            return (hf <= h <= ht) if hf <= ht else (h >= hf or h <= ht)
        events = [e for e in events if e.get("_event_dt") and _in(e["_event_dt"].hour)]
    # Layer 3: list-type filter
    lt_filter = set((cfg.filters or {}).get("list_types") or [])
    if lt_filter:
        events = [e for e in events
                  if (e.get("_list_type") or "").lower() in lt_filter]
    agg = _aggregate(events, cfg.metric, U, cfg.dim_x, cfg.dim_y, cfg.output)

    # Compare-to delta (Layer 4)
    compare_block = None
    if cfg.compare_to_previous and cfg.window in ("last_7d", "last_30d", "last_90d", "last_180d"):
        days = {"last_7d": 7, "last_30d": 30, "last_90d": 90, "last_180d": 180}[cfg.window]
        prev_to = win[0]
        prev_from = prev_to - timedelta(days=days)
        prev_U = U  # universe doesn't change
        prev_events = await METRIC_FNS[cfg.metric](prev_U, (prev_from, prev_to), cfg.filters or {})
        prev_agg = _aggregate(prev_events, cfg.metric, prev_U, cfg.dim_x, cfg.dim_y, {})
        compare_block = {
            "from": prev_from.isoformat(),
            "to": prev_to.isoformat(),
            "result": prev_agg,
        }

    return {
        "config": cfg.dict(),
        "window": {"from": win[0].isoformat() if win[0] else None,
                   "to": win[1].isoformat() if win[1] else None},
        "asset_pool_size": len(U["assets"]),
        "event_count": len(events),
        "compare_to": compare_block,
        **agg,
        "generated_at": now_ist().isoformat(),
    }


@router.post("/api/reports/builder/run/{user_id}")
async def builder_run(user_id: str, cfg: BuilderConfig):
    await _ensure_sa(user_id)
    result = await _run_config(cfg)
    # Layer 6 — log every run for history (lightweight, no result blob)
    try:
        await report_runs_collection.insert_one({
            "user_id": user_id,
            "config": cfg.dict(),
            "row_count": len(result.get("rows", [])) if "rows" in result else len(result.get("row_keys", [])),
            "event_count": result.get("event_count"),
            "asset_pool_size": result.get("asset_pool_size"),
            "created_at": now_ist().isoformat(),
        })
    except Exception:
        pass  # logging failure must not break the user's run
    return result


@router.get("/api/reports/builder/runs/{user_id}")
async def list_runs(user_id: str, limit: int = Query(20, ge=1, le=200)):
    """Layer 6 — most recent runs for this user."""
    await _ensure_sa(user_id)
    docs = await report_runs_collection.find({"user_id": user_id}).sort("created_at", -1).limit(limit).to_list(limit)
    return [{
        "_id": str(d["_id"]),
        "config": d.get("config"),
        "row_count": d.get("row_count"),
        "event_count": d.get("event_count"),
        "created_at": d.get("created_at"),
    } for d in docs]


# ════════════════════════════════════════════════════════════════════════════
# Saved reports CRUD
# ════════════════════════════════════════════════════════════════════════════
class SavedReportCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=120)
    config: BuilderConfig


@router.get("/api/reports/builder/saved/{user_id}")
async def list_saved(user_id: str):
    await _ensure_sa(user_id)
    docs = await saved_reports_collection.find({"owner_id": user_id}).sort("created_at", -1).to_list(500)
    return [{
        "_id": str(d["_id"]),
        "name": d["name"],
        "config": d["config"],
        "created_at": d.get("created_at"),
    } for d in docs]


@router.post("/api/reports/builder/save/{user_id}")
async def save_report(user_id: str, payload: SavedReportCreate):
    await _ensure_sa(user_id)
    doc = {
        "owner_id": user_id,
        "name": payload.name,
        "config": payload.config.dict(),
        "created_at": now_ist().isoformat(),
    }
    res = await saved_reports_collection.insert_one(doc)
    return {"_id": str(res.inserted_id), **{k: v for k, v in doc.items() if k != "_id"}}


@router.delete("/api/reports/builder/saved/{report_id}/{user_id}")
async def delete_saved(report_id: str, user_id: str):
    await _ensure_sa(user_id)
    try:
        oid = ObjectId(report_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    res = await saved_reports_collection.delete_one({"_id": oid, "owner_id": user_id})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Not found")
    return {"deleted": True}


# ════════════════════════════════════════════════════════════════════════════
# Featured library — 8 ready-made configs (no DB hit, just static list)
# ════════════════════════════════════════════════════════════════════════════
FEATURED: List[dict] = [
    {
        "id": "today_new_defects_by_station",
        "name": "Today's New Defects by Station",
        "description": "Count of new OL entries opened in the last 24 hours, grouped by station.",
        "config": {"metric": "defect_frequency", "dim_x": "station",
                   "window": "last_7d", "filters": {}},
    },
    {
        "id": "weekly_mttr_by_asset_type",
        "name": "Weekly MTTR by Asset Type",
        "description": "Median repair hours over the last 7 days, grouped by asset type.",
        "config": {"metric": "mttr", "dim_x": "asset_type",
                   "window": "last_7d", "filters": {}},
    },
    {
        "id": "worst_locations_30d",
        "name": "Worst Locations (last 30d)",
        "description": "% Working by Location — worst-first.",
        "config": {"metric": "pct_working", "dim_x": "location",
                   "window": "last_30d", "filters": {}},
    },
    {
        "id": "slow_supervisors_30d",
        "name": "Slowest Supervisors by MTTR (last 30d)",
        "description": "Median repair hours per supervisor.",
        "config": {"metric": "mttr", "dim_x": "supervisor",
                   "window": "last_30d", "filters": {}},
    },
    {
        "id": "defect_heatmap_station_x_type",
        "name": "Defect Frequency Heatmap — Station × Asset Type",
        "description": "Count of defects per station × asset type. (cross-tab)",
        "config": {"metric": "defect_frequency", "dim_x": "station",
                   "dim_y": "asset_type", "window": "last_30d", "filters": {}},
    },
    {
        "id": "rejection_rate_supervisor_90d",
        "name": "Rejection Rate by Supervisor (last 90d)",
        "description": "% of mark-working claims rejected per supervisor.",
        "config": {"metric": "rejection_rate", "dim_x": "supervisor",
                   "window": "last_90d", "filters": {}},
    },
    {
        "id": "fy_mttr_by_dept",
        "name": "FY-to-date MTTR by Department",
        "description": "Median repair hours since Apr 1 per department.",
        "config": {"metric": "mttr", "dim_x": "department",
                   "window": "fy", "filters": {}},
    },
    {
        "id": "inspection_volume_inspector_7d",
        "name": "Inspection Volume by Inspector (last 7d)",
        "description": "Inspection items per inspector over the last week.",
        "config": {"metric": "inspection_volume", "dim_x": "supervisor",
                   "window": "last_7d", "filters": {}},
    },
]


@router.get("/api/reports/builder/featured")
async def featured():
    return FEATURED


# ════════════════════════════════════════════════════════════════════════════
# Dimensions metadata (for filter dropdowns)
# ════════════════════════════════════════════════════════════════════════════
@router.get("/api/reports/builder/dimensions/{user_id}")
async def dims_meta(user_id: str):
    await _ensure_sa(user_id)
    stations = await stations_collection.find({}).to_list(2000)
    depts = await departments_collection.find({}).to_list(500)
    types = await asset_types_collection.find({}).to_list(2000)
    return {
        "stations": [{"id": str(s["_id"]), "name": s.get("name", "")} for s in stations],
        "departments": [{"id": str(d["_id"]), "name": d.get("name", "")} for d in depts],
        "asset_types": [{"id": str(t["_id"]), "name": t.get("name", "")} for t in types],
        "metrics": [
            {"id": "pct_working", "name": "% Working", "unit": "%", "kind": "pct"},
            {"id": "mttr", "name": "MTTR (median repair hours)", "unit": "hrs", "kind": "hrs"},
            {"id": "defect_frequency", "name": "Defect Frequency (new defects)", "unit": "count", "kind": "count"},
            {"id": "rejection_rate", "name": "Rejection Rate", "unit": "%", "kind": "pct"},
            {"id": "inspection_volume", "name": "Inspection Volume", "unit": "count", "kind": "count"},
            {"id": "first_time_fix_rate", "name": "First-time-fix Rate", "unit": "%", "kind": "pct"},
            {"id": "recurrence_rate", "name": "Recurrence Rate (re-defected)", "unit": "%", "kind": "pct"},
            {"id": "backlog_age", "name": "Backlog Age (currently-open hrs)", "unit": "hrs", "kind": "hrs"},
            {"id": "throughput", "name": "Throughput (resolutions)", "unit": "count", "kind": "count"},
            {"id": "avg_approval_lag", "name": "Avg Approval Lag (mark→approve)", "unit": "hrs", "kind": "hrs"},
            {"id": "pct_pending", "name": "% Pending Verification", "unit": "%", "kind": "pct"},
            {"id": "inspection_coverage", "name": "Inspection Coverage", "unit": "%", "kind": "pct"},
        ],
        "dimensions": [
            {"id": "station", "name": "Station"},
            {"id": "location", "name": "Location"},
            {"id": "department", "name": "Department"},
            {"id": "asset_type", "name": "Asset Type"},
            {"id": "asset", "name": "Per Asset"},
            {"id": "supervisor", "name": "Person (supervisor / actor)"},
            {"id": "ro", "name": "Reporting Officer (asset's RO)"},
            {"id": "asup", "name": "Approving Supervisor"},
            {"id": "inspector", "name": "Inspector (event)"},
            {"id": "reporter", "name": "Reporter (who opened defect)"},
            {"id": "list_type", "name": "List Type (orange/red/yellow)"},
            {"id": "defect_age_band", "name": "Defect Age Band (0-2h / 2-8h / …)"},
            {"id": "repair_age_band", "name": "Repair Age Band"},
            {"id": "hour_of_day", "name": "Hour of Day"},
            {"id": "day_of_week", "name": "Day of Week"},
            {"id": "time_bucket_day", "name": "Day"},
            {"id": "time_bucket_week", "name": "Week"},
            {"id": "time_bucket_month", "name": "Month"},
            {"id": "time_bucket_quarter", "name": "Quarter (FY-aligned)"},
        ],
        "windows": [
            {"id": "last_7d", "name": "Last 7 days"},
            {"id": "last_30d", "name": "Last 30 days"},
            {"id": "last_90d", "name": "Last 90 days"},
            {"id": "last_180d", "name": "Last 180 days"},
            {"id": "fy", "name": "Current Financial Year"},
            {"id": "all_time", "name": "All time"},
            {"id": "custom", "name": "Custom range"},
        ],
        "asset_statuses": ["working", "pending_approval", "defective"],
        "list_types": ["orange", "red", "yellow"],
        "viz_options": ["table", "bar", "stacked_bar", "donut", "heatmap", "line"],
    }


# ════════════════════════════════════════════════════════════════════════════
# Exports — CSV / Excel / PDF
# ════════════════════════════════════════════════════════════════════════════
def _flatten_for_export(result: dict) -> tuple:
    """Return (headers, rows) flattening either single-dim rows or 2-dim matrix.
    Hours-style metrics (mttr/backlog_age/avg_approval_lag) expand percentile cols."""
    cfg = result["config"]
    metric = cfg["metric"]
    extra_cols = []
    if metric in {"mttr", "backlog_age", "avg_approval_lag"}:
        extra_cols = ["p25", "p75", "p90", "p99", "min", "max", "mean"]

    if "matrix" in result:
        headers = ["Row \\ Col"] + result["col_labels"]
        rows = []
        for ridx, rk in enumerate(result["row_keys"]):
            row = [result["row_labels"][ridx]]
            for cell in result["matrix"][ridx]:
                v = cell.get("value")
                row.append("" if v is None else v)
            rows.append(row)
        return headers, rows

    headers = [_dim_label(cfg["dim_x"]), "Value", "n"] + [c.upper() for c in extra_cols]
    rows = []
    for r in result["rows"]:
        row = [r["label_x"], r["value"], r["n"]]
        for c in extra_cols:
            row.append(r.get(c))
        rows.append(row)
    return headers, rows


def _dim_label(dim: str) -> str:
    return {
        "station": "Station", "location": "Location", "department": "Department",
        "asset_type": "Asset Type", "asset": "Asset", "supervisor": "Person",
        "ro": "Reporting Officer", "asup": "Approving Supervisor",
        "inspector": "Inspector", "reporter": "Reporter",
        "list_type": "List Type",
        "defect_age_band": "Defect Age", "repair_age_band": "Repair Age",
        "hour_of_day": "Hour", "day_of_week": "Weekday",
        "time_bucket_day": "Day", "time_bucket_week": "Week",
        "time_bucket_month": "Month", "time_bucket_quarter": "Quarter",
    }.get(dim, dim)


def _metric_label(m: str) -> str:
    return {
        "pct_working": "% Working",
        "mttr": "MTTR (median hrs)",
        "defect_frequency": "Defect Frequency",
        "rejection_rate": "Rejection Rate (%)",
        "inspection_volume": "Inspection Volume",
        "first_time_fix_rate": "First-time-fix Rate (%)",
        "recurrence_rate": "Recurrence Rate (%)",
        "backlog_age": "Backlog Age (hrs)",
        "throughput": "Throughput (resolutions)",
        "avg_approval_lag": "Avg Approval Lag (hrs)",
        "pct_pending": "% Pending Verification",
        "inspection_coverage": "Inspection Coverage (%)",
    }.get(m, m)


@router.post("/api/reports/builder/export/csv/{user_id}")
async def export_csv(user_id: str, cfg: BuilderConfig):
    await _ensure_sa(user_id)
    result = await _run_config(cfg)
    headers, rows = _flatten_for_export(result)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"Metric: {_metric_label(cfg.metric)}"])
    w.writerow([f"Window: {cfg.window}", f"Generated: {result['generated_at']}"])
    w.writerow([])
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    out = io.BytesIO(buf.getvalue().encode("utf-8"))
    fname = f"builder-{cfg.metric}-{now_ist().strftime('%Y%m%d-%H%M')}.csv"
    return StreamingResponse(out, media_type="text/csv",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.post("/api/reports/builder/export/excel/{user_id}")
async def export_excel(user_id: str, cfg: BuilderConfig):
    await _ensure_sa(user_id)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    result = await _run_config(cfg)
    headers, rows = _flatten_for_export(result)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    teal = PatternFill("solid", fgColor="0E7C6B")
    bold_white = Font(bold=True, color="FFFFFF")
    ws.append([f"Metric: {_metric_label(cfg.metric)}"])
    ws.append([f"Window: {cfg.window}", f"Generated: {result['generated_at']}"])
    ws.append([])
    ws.append(headers)
    for c in ws[4]:
        c.fill = teal; c.font = bold_white
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A5"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 22
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    fname = f"builder-{cfg.metric}-{now_ist().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/api/reports/builder/export/pdf/{user_id}")
async def export_pdf(user_id: str, cfg: BuilderConfig):
    await _ensure_sa(user_id)
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    result = await _run_config(cfg)
    headers, rows = _flatten_for_export(result)
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title="Custom Report")
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Heading1"], fontSize=14,
                           textColor=colors.HexColor("#0e7c6b"), spaceAfter=4)
    sub = ParagraphStyle("s", parent=styles["Normal"], fontSize=8,
                         textColor=colors.HexColor("#64748b"))
    elements = [Paragraph(_metric_label(cfg.metric), title)]
    elements.append(Paragraph(
        f"Dim X: <b>{_dim_label(cfg.dim_x)}</b>"
        + (f" · Dim Y: <b>{_dim_label(cfg.dim_y)}</b>" if cfg.dim_y else "")
        + f" · Window: <b>{cfg.window}</b> · {len(rows)} row(s)", sub))
    elements.append(Spacer(1, 8))
    if rows:
        data = [headers] + [[("" if c is None else str(c)) for c in r] for r in rows]
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7c6b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f8fafc")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No data for this configuration.", styles["Normal"]))
    doc.build(elements)
    out.seek(0)
    fname = f"builder-{cfg.metric}-{now_ist().strftime('%Y%m%d-%H%M')}.pdf"
    return StreamingResponse(out, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ════════════════════════════════════════════════════════════════════════════
# Layer 5 — Multi-section dossiers
# ════════════════════════════════════════════════════════════════════════════
class DossierSection(BaseModel):
    title: str
    config: BuilderConfig


class DossierConfig(BaseModel):
    title: str = "Custom Report"
    subtitle: Optional[str] = None
    cover: Optional[Dict[str, Any]] = None      # logo_text, prepared_for, prepared_by, footer
    sections: List[DossierSection]


class DossierCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=120)
    dossier: DossierConfig


@router.post("/api/reports/builder/dossier/run/{user_id}")
async def dossier_run(user_id: str, dossier: DossierConfig):
    """Run every section, return an array of results."""
    await _ensure_sa(user_id)
    out = []
    for sec in dossier.sections:
        try:
            r = await _run_config(sec.config)
            out.append({"title": sec.title, "result": r})
        except HTTPException as e:
            out.append({"title": sec.title, "error": e.detail})
    return {
        "title": dossier.title,
        "subtitle": dossier.subtitle,
        "cover": dossier.cover,
        "sections": out,
        "generated_at": now_ist().isoformat(),
    }


@router.get("/api/reports/builder/dossier/saved/{user_id}")
async def list_dossiers(user_id: str):
    await _ensure_sa(user_id)
    docs = await saved_dossiers_collection.find({"owner_id": user_id}).sort("created_at", -1).to_list(500)
    return [{
        "_id": str(d["_id"]), "name": d["name"], "dossier": d["dossier"],
        "created_at": d.get("created_at"),
    } for d in docs]


@router.post("/api/reports/builder/dossier/save/{user_id}")
async def save_dossier(user_id: str, payload: DossierCreate):
    await _ensure_sa(user_id)
    doc = {
        "owner_id": user_id,
        "name": payload.name,
        "dossier": payload.dossier.dict(),
        "created_at": now_ist().isoformat(),
    }
    res = await saved_dossiers_collection.insert_one(doc)
    return {"_id": str(res.inserted_id), **{k: v for k, v in doc.items() if k != "_id"}}


@router.delete("/api/reports/builder/dossier/saved/{dossier_id}/{user_id}")
async def delete_dossier(dossier_id: str, user_id: str):
    await _ensure_sa(user_id)
    try:
        oid = ObjectId(dossier_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    res = await saved_dossiers_collection.delete_one({"_id": oid, "owner_id": user_id})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Not found")
    return {"deleted": True}


@router.post("/api/reports/builder/dossier/export/pdf/{user_id}")
async def dossier_export_pdf(user_id: str, dossier: DossierConfig):
    """Generate a multi-page PDF with cover + each section."""
    await _ensure_sa(user_id)
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate, NextPageTemplate,
        Paragraph, Spacer, Table, TableStyle, PageBreak,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    out = io.BytesIO()
    portrait_frame = Frame(15*mm, 15*mm, A4[0] - 30*mm, A4[1] - 30*mm, id="portrait")
    landsize = landscape(A4)
    landscape_frame = Frame(10*mm, 10*mm, landsize[0] - 20*mm, landsize[1] - 20*mm, id="landscape")
    doc = BaseDocTemplate(out, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                          topMargin=15*mm, bottomMargin=15*mm,
                          title=dossier.title)
    doc.addPageTemplates([
        PageTemplate(id="portrait", frames=[portrait_frame], pagesize=A4),
        PageTemplate(id="landscape", frames=[landscape_frame], pagesize=landsize),
    ])
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=24,
                                 textColor=rl_colors.HexColor("#0e7c6b"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=14,
                        textColor=rl_colors.HexColor("#0e7c6b"))
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10,
                         textColor=rl_colors.HexColor("#64748b"))
    cell_small = ParagraphStyle("c", parent=styles["Normal"], fontSize=8)

    elements = []
    # Cover page
    cover = dossier.cover or {}
    elements.append(Spacer(1, 60))
    elements.append(Paragraph(dossier.title, title_style))
    elements.append(Spacer(1, 8))
    if dossier.subtitle:
        elements.append(Paragraph(dossier.subtitle, sub))
    elements.append(Spacer(1, 30))
    if cover.get("prepared_for"):
        elements.append(Paragraph(f"<b>Prepared for:</b> {cover['prepared_for']}", styles["Normal"]))
    if cover.get("prepared_by"):
        elements.append(Paragraph(f"<b>Prepared by:</b> {cover['prepared_by']}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Generated:</b> {now_ist().strftime('%d %b %Y, %H:%M IST')}", styles["Normal"]))
    if cover.get("footer"):
        elements.append(Spacer(1, 40))
        elements.append(Paragraph(cover["footer"], sub))

    # Sections
    for sec in dossier.sections:
        elements.append(PageBreak())
        elements.append(Paragraph(sec.title, h2))
        try:
            r = await _run_config(sec.config)
        except HTTPException as e:
            elements.append(Paragraph(f"Error: {e.detail}", styles["Normal"]))
            continue
        meta_line = (f"{_metric_label(sec.config.metric)} · {_dim_label(sec.config.dim_x)}"
                     + (f" × {_dim_label(sec.config.dim_y)}" if sec.config.dim_y else "")
                     + f" · Window: {sec.config.window} · n={r.get('event_count', 0)}")
        elements.append(Paragraph(meta_line, sub))
        # Annotation if present
        ann = (sec.config.annotations or {}).get("note")
        if ann:
            elements.append(Spacer(1, 4))
            elements.append(Paragraph(f"<i>{ann}</i>", cell_small))
        elements.append(Spacer(1, 8))
        headers, rows = _flatten_for_export(r)
        if rows:
            data = [headers] + [[("" if c is None else str(c)) for c in row] for row in rows[:60]]
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#0e7c6b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [rl_colors.white, rl_colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            elements.append(t)
            if len(rows) > 60:
                elements.append(Spacer(1, 4))
                elements.append(Paragraph(f"… {len(rows) - 60} more rows truncated for PDF.", cell_small))
        else:
            elements.append(Paragraph("No data.", styles["Normal"]))

    doc.build(elements)
    out.seek(0)
    fname = f"dossier-{now_ist().strftime('%Y%m%d-%H%M')}.pdf"
    return StreamingResponse(out, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.post("/api/reports/builder/dossier/export/excel/{user_id}")
async def dossier_export_excel(user_id: str, dossier: DossierConfig):
    """One workbook with one sheet per section."""
    await _ensure_sa(user_id)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.append([dossier.title])
    cover["A1"].font = Font(size=20, bold=True, color="0E7C6B")
    if dossier.subtitle:
        cover.append([dossier.subtitle])
    cover.append([])
    cover.append([f"Generated: {now_ist().isoformat()}"])
    if dossier.cover:
        for k, v in (dossier.cover or {}).items():
            cover.append([k.replace("_", " ").title(), str(v)])
    teal = PatternFill("solid", fgColor="0E7C6B")
    bold_white = Font(bold=True, color="FFFFFF")
    for i, sec in enumerate(dossier.sections, 1):
        try:
            r = await _run_config(sec.config)
        except HTTPException:
            continue
        ws = wb.create_sheet(f"{i}. {sec.title[:25]}")
        ws.append([sec.title]); ws["A1"].font = Font(bold=True, size=14, color="0E7C6B")
        ws.append([f"{_metric_label(sec.config.metric)} · {_dim_label(sec.config.dim_x)}"
                   + (f" × {_dim_label(sec.config.dim_y)}" if sec.config.dim_y else "")
                   + f" · {sec.config.window}"])
        ws.append([])
        headers, rows = _flatten_for_export(r)
        ws.append(headers)
        for c in ws[4]:
            c.fill = teal; c.font = bold_white
            c.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A5"
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 22
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    fname = f"dossier-{now_ist().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

