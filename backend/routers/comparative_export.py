"""
Comparative Reports — PDF/Excel export with configurable sections.

Endpoints:
  POST /api/reports/comparative/export/pdf/{user_id}
  POST /api/reports/comparative/export/excel/{user_id}

Body (both):
  {
    "window_days": "7|15|30|90|fy|all" (default "90"),
    "stat":        "median|mean" (default "median"),
    "dept_id":     "<id>" (optional),
    "asset_type_ids": ["<id>", ...] (optional, falls back to top-5 like UI),
    "drill_state": {"level": "station|location_summary|location_types|asset",
                    "parent_id": "<id|null>",
                    "parent_asset_type_id": "<id|null>"} (default station/null/null),
    "sections": {
      "card_a":            true,    # MTTR by Asset Type
      "card_b":            true,    # Peer comparison (radar matrix)
      "card_c_current":    true,    # Current drill state from the UI
      "card_c_full":       false,   # Full 4-level hierarchy
      "defective":         true,    # Defective-only appendix (Orange/Red list)
      "remarks":           true,    # Per-asset remarks (last 5)
      "last_inspection":   true     # Last inspection per asset
    },
    "style":   "detailed|compact"  (PDF only, default detailed)
  }
"""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from bson import ObjectId
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from database import (
    users_collection, assets_collection, asset_types_collection,
    stations_collection, locations_collection, departments_collection,
    orange_list_collection, inspections_collection, remarks_collection,
    now_ist,
)
from routers.comparative import (
    _window_from_days, _percentile, _hrs_stats, _resolved_repair_hours,
    _parse_dt, _user_station_ids, PALETTE,
)

# Reportlab imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether,
)
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Circle, Polygon
from reportlab.graphics import renderPDF

# Openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()


# ─── Request model ────────────────────────────────────────────────────────
class DrillState(BaseModel):
    level: str = "station"
    parent_id: Optional[str] = None
    parent_asset_type_id: Optional[str] = None


class Sections(BaseModel):
    card_a: bool = True
    card_b: bool = True
    card_c_current: bool = True
    card_c_full: bool = False
    defective: bool = True
    remarks: bool = True
    last_inspection: bool = True


class ExportRequest(BaseModel):
    window_days: str = "90"
    stat: str = "median"
    dept_id: Optional[str] = None
    asset_type_ids: Optional[List[str]] = None
    drill_state: DrillState = DrillState()
    sections: Sections = Sections()
    style: str = "detailed"


# ─── Helpers — datasets used by both PDF and Excel ────────────────────────
async def _resolve_asset_types(req: ExportRequest, types_in_scope_ids: set,
                               type_by_id: Dict[str, dict],
                               win, f_dt, t_dt) -> List[str]:
    """Mirror UI default — top-5 by event count if none supplied."""
    explicit = [t for t in (req.asset_type_ids or []) if t in types_in_scope_ids]
    if explicit:
        return explicit
    all_ols = await orange_list_collection.find({"status": "resolved"}).to_list(50000)
    all_assets = await assets_collection.find({}).to_list(20000)
    type_of_asset = {str(a["_id"]): a.get("asset_type_id") for a in all_assets}
    type_counts: Dict[str, int] = {}
    for ol in all_ols:
        mw = _parse_dt(ol.get("marked_working_at"))
        if not mw:
            continue
        if f_dt and mw < f_dt:
            continue
        if t_dt and mw > t_dt:
            continue
        tid = type_of_asset.get(ol.get("asset_id"))
        if tid and tid in types_in_scope_ids:
            type_counts[tid] = type_counts.get(tid, 0) + 1
    out = [t for t, _ in sorted(type_counts.items(), key=lambda kv: -kv[1])[:5]]
    return out or list(types_in_scope_ids)[:5]


async def _build_data_bundle(user_id: str, req: ExportRequest) -> Dict[str, Any]:
    """Pre-fetches all data needed by every section, once."""
    try:
        user = await users_collection.find_one({"_id": ObjectId(user_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user_id")
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    f_dt, t_dt = _window_from_days(req.window_days)
    win = (f_dt, t_dt)

    types = await asset_types_collection.find({}).to_list(2000)
    type_by_id = {str(t["_id"]): t for t in types}
    departments = await departments_collection.find({}).to_list(200)
    dept_by_id = {str(d["_id"]): d for d in departments}

    types_in_scope_ids = {
        str(t["_id"]) for t in types
        if not req.dept_id or t.get("department_id") == req.dept_id
    }
    selected_type_ids = await _resolve_asset_types(req, types_in_scope_ids,
                                                   type_by_id, win, f_dt, t_dt)
    type_palette = {tid: PALETTE[i % len(PALETTE)] for i, tid in enumerate(selected_type_ids)}

    user_stns = set(_user_station_ids(user))
    asset_q: Dict[str, Any] = {"asset_type_id": {"$in": selected_type_ids}}
    if user_stns:
        asset_q["station_id"] = {"$in": list(user_stns)}
    assets = await assets_collection.find(asset_q).to_list(20000)
    asset_by_id = {str(a["_id"]): a for a in assets}
    asset_ids = list(asset_by_id.keys())

    stations = await stations_collection.find({}).to_list(2000)
    station_by_id = {str(s["_id"]): s for s in stations}
    locations = await locations_collection.find({}).to_list(5000)
    location_by_id = {str(loc["_id"]): loc for loc in locations}

    ols_resolved = await orange_list_collection.find(
        {"asset_id": {"$in": asset_ids}, "status": "resolved"}).to_list(50000)
    ols_open = await orange_list_collection.find(
        {"asset_id": {"$in": asset_ids}, "status": {"$ne": "resolved"}}).to_list(50000)

    # Last inspections per asset (only fetch if needed)
    last_inspections: Dict[str, dict] = {}
    if req.sections.last_inspection or req.sections.defective:
        ins_cur = inspections_collection.find(
            {"asset_id": {"$in": asset_ids}}, sort=[("inspection_at", -1)]).limit(20000)
        async for ins in ins_cur:
            aid = str(ins.get("asset_id"))
            if aid not in last_inspections:
                last_inspections[aid] = ins

    # Remarks per asset (only fetch if needed) — last 5 each
    remarks_by_asset: Dict[str, List[dict]] = {}
    if req.sections.remarks:
        # Remarks are linked to OL entries via orange_list_id; collect all OL ids for our assets
        all_ol_ids = [str(ol["_id"]) for ol in (ols_resolved + ols_open)]
        if all_ol_ids:
            rem_cur = remarks_collection.find(
                {"orange_list_id": {"$in": all_ol_ids}},
                sort=[("created_at", -1)]).limit(20000)
            ol_to_asset = {str(ol["_id"]): str(ol.get("asset_id"))
                           for ol in (ols_resolved + ols_open)}
            async for rem in rem_cur:
                ol_id = str(rem.get("orange_list_id"))
                aid = ol_to_asset.get(ol_id)
                if not aid:
                    continue
                lst = remarks_by_asset.setdefault(aid, [])
                if len(lst) < 5:
                    lst.append(rem)

    return dict(
        user=user, req=req, f_dt=f_dt, t_dt=t_dt, win=win,
        type_by_id=type_by_id, dept_by_id=dept_by_id,
        types_in_scope_ids=types_in_scope_ids,
        selected_type_ids=selected_type_ids,
        type_palette=type_palette,
        asset_by_id=asset_by_id, asset_ids=asset_ids,
        station_by_id=station_by_id, location_by_id=location_by_id,
        ols_resolved=ols_resolved, ols_open=ols_open,
        last_inspections=last_inspections,
        remarks_by_asset=remarks_by_asset,
    )


def _semantic_color(value, peer_max):
    """Low MTTR = green, high MTTR = red. Returns reportlab Color or None."""
    if value is None or peer_max is None or peer_max == 0:
        return colors.HexColor("#94a3b8")
    t = max(0.0, min(1.0, value / peer_max))
    if t < 0.33:
        return colors.HexColor("#10b981")
    if t < 0.66:
        return colors.HexColor("#eab308")
    if t < 0.85:
        return colors.HexColor("#f97316")
    return colors.HexColor("#dc2626")


def _section_a_rows(bundle: Dict[str, Any]) -> List[List[Any]]:
    """Card A — MTTR by asset-type rows."""
    req: ExportRequest = bundle["req"]
    type_by_id = bundle["type_by_id"]
    by_type: Dict[str, set] = {}
    for a in bundle["asset_by_id"].values():
        tid = a.get("asset_type_id")
        if tid in bundle["selected_type_ids"]:
            by_type.setdefault(tid, set()).add(str(a["_id"]))
    rows = []
    for tid, ids in by_type.items():
        hours = _resolved_repair_hours(bundle["ols_resolved"], ids, bundle["win"])
        s = _hrs_stats(hours)
        rows.append({
            "asset_type_id": tid,
            "label": (type_by_id.get(tid) or {}).get("name", "—"),
            **s,
        })
    rows = [r for r in rows if r.get("n", 0) > 0]
    rows.sort(key=lambda r: -(r.get(req.stat) or 0))
    return rows


def _section_b_matrix(bundle: Dict[str, Any]) -> Dict[str, Any]:
    """Card B — peer-supervisor radar matrix."""
    req: ExportRequest = bundle["req"]
    user = bundle["user"]
    dept_id = req.dept_id or user.get("department_id")
    return {
        "axes": [{"id": tid, "name": (bundle["type_by_id"].get(tid) or {}).get("name", "—")}
                 for tid in bundle["selected_type_ids"]],
        "dept_id": dept_id,
        "_needs_compute": True,  # caller will use existing endpoint logic
    }


def _section_defective_rows(bundle: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Currently-open defects in scope."""
    out = []
    now = now_ist()
    for ol in bundle["ols_open"]:
        aid = str(ol.get("asset_id"))
        a = bundle["asset_by_id"].get(aid)
        if not a:
            continue
        ds = _parse_dt(ol.get("defective_since"))
        days_open = round((now - ds).total_seconds() / 86400, 1) if ds else None
        st = bundle["station_by_id"].get(a.get("station_id")) or {}
        loc = bundle["location_by_id"].get(a.get("location_id")) or {}
        tp = bundle["type_by_id"].get(a.get("asset_type_id")) or {}
        out.append({
            "asset_id": aid,
            "asset_number": a.get("asset_number"),
            "asset_type": tp.get("name", "—"),
            "station": st.get("name", "—"),
            "location": loc.get("name", "—"),
            "list_type": (ol.get("list_type") or "").lower(),
            "status": ol.get("status"),
            "defective_since": str(ol.get("defective_since") or ""),
            "days_open": days_open,
        })
    out.sort(key=lambda r: -(r.get("days_open") or 0))
    return out


# ─── PDF ──────────────────────────────────────────────────────────────────
def _draw_horizontal_bar(value: float, peer_max: float, width_mm: float, height_mm: float,
                         color_hex: str = None) -> Drawing:
    """Returns a small ReportLab Drawing with a colored bar (proxy for cylinder)."""
    w = width_mm * mm
    h = height_mm * mm
    d = Drawing(w, h)
    if peer_max <= 0:
        peer_max = 1
    if value is None:
        return d
    bw = max(2, (value / peer_max) * w)
    fill = colors.HexColor(color_hex) if color_hex else _semantic_color(value, peer_max)
    # Cylinder-ish: rounded rect via 2 ellipses + body
    d.add(Rect(0, 0, bw, h, fillColor=fill, strokeColor=fill, rx=h/2, ry=h/2))
    d.add(Rect(0, h*0.55, bw, h*0.25,
               fillColor=colors.HexColor("#ffffff"), strokeColor=None,
               fillOpacity=0.35))
    return d


def _draw_radar(axes: List[Dict[str, str]], series: List[Dict[str, Any]], size_mm: float = 90) -> Drawing:
    """Compact radar chart for embedding in PDF."""
    sz = size_mm * mm
    d = Drawing(sz, sz)
    cx, cy = sz / 2, sz / 2
    r_max = sz / 2 * 0.78
    n = len(axes)
    if n < 3:
        return d
    import math
    g_max = 0
    for s in series:
        for v in s.get("values", []):
            if v.get("value") is not None and v["value"] > g_max:
                g_max = v["value"]
    if g_max <= 0:
        g_max = 1
    angle = lambda i: -math.pi / 2 + (i * 2 * math.pi) / n
    pt = lambda i, frac: (cx + r_max * frac * math.cos(angle(i)),
                          cy + r_max * frac * math.sin(angle(i)))

    # Grid rings
    for lv in (0.25, 0.5, 0.75, 1.0):
        pts = []
        for i in range(n):
            x, y = pt(i, lv)
            pts.extend([x, y])
        d.add(Polygon(pts, fillColor=None,
                      strokeColor=colors.HexColor("#cbd5e1"),
                      strokeDashArray=[2, 3], strokeWidth=0.4))
    # Axes
    for i in range(n):
        x, y = pt(i, 1)
        d.add(Line(cx, cy, x, y, strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.4))

    # Series — peers first then self
    sorted_series = sorted(series, key=lambda s: 1 if s.get("is_self") else 0)
    for s in sorted_series:
        is_self = s.get("is_self")
        stroke = colors.HexColor("#0e7c6b") if is_self else colors.HexColor("#60a5fa")
        pts = []
        for i, v in enumerate(s.get("values", [])):
            frac = (v.get("value") or 0) / g_max
            x, y = pt(i, frac)
            pts.extend([x, y])
        if not pts:
            continue
        d.add(Polygon(pts, fillColor=stroke, strokeColor=stroke,
                      fillOpacity=0.35 if is_self else 0.12,
                      strokeWidth=1.6 if is_self else 0.9))

    # Axis labels
    for i, ax in enumerate(axes):
        x, y = pt(i, 1.10)
        align = "middle" if abs(x - cx) < 4 else ("start" if x > cx else "end")
        nm = ax["name"]
        if len(nm) > 14:
            nm = nm[:13] + "…"
        d.add(String(x, y, nm, fontSize=6.5, textAnchor=align,
                     fillColor=colors.HexColor("#0f172a")))
    return d


def _filter_chips_para(bundle, styles):
    req: ExportRequest = bundle["req"]
    parts = [f"<b>Window</b>: {req.window_days}", f"<b>Stat</b>: {req.stat}"]
    if req.dept_id:
        d = bundle["dept_by_id"].get(req.dept_id) or {}
        parts.append(f"<b>Department</b>: {d.get('name', '—')}")
    if bundle["selected_type_ids"]:
        names = [(bundle["type_by_id"].get(t) or {}).get("name", "—")
                 for t in bundle["selected_type_ids"]]
        parts.append(f"<b>Asset Types</b>: {', '.join(names)}")
    parts.append(f"<b>Generated</b>: {now_ist().strftime('%d %b %Y, %H:%M IST')}")
    return Paragraph(" &nbsp; · &nbsp; ".join(parts), styles["small"])


async def _build_pdf(user_id: str, req: ExportRequest) -> bytes:
    bundle = await _build_data_bundle(user_id, req)
    user = bundle["user"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm,
                            title="Comparative Reports")
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle("title", parent=base["Title"], fontSize=20,
                                textColor=colors.HexColor("#0e7c6b"),
                                alignment=0, spaceAfter=4),
        "h1": ParagraphStyle("h1", parent=base["Heading1"], fontSize=14,
                             textColor=colors.HexColor("#0f172a"), spaceBefore=8, spaceAfter=4),
        "h2": ParagraphStyle("h2", parent=base["Heading2"], fontSize=11,
                             textColor=colors.HexColor("#334155"), spaceBefore=6, spaceAfter=2),
        "body": ParagraphStyle("body", parent=base["BodyText"], fontSize=9,
                               textColor=colors.HexColor("#0f172a"), leading=12),
        "small": ParagraphStyle("small", parent=base["BodyText"], fontSize=8,
                                textColor=colors.HexColor("#475569"), leading=11),
        "muted": ParagraphStyle("muted", parent=base["BodyText"], fontSize=8,
                                textColor=colors.HexColor("#94a3b8"), leading=10),
        "cell": ParagraphStyle("cell", parent=base["BodyText"], fontSize=7.5,
                               textColor=colors.HexColor("#0f172a"),
                               leading=9.5, wordWrap="CJK"),
        "cell_b": ParagraphStyle("cell_b", parent=base["BodyText"], fontSize=7.5,
                                 textColor=colors.HexColor("#0f172a"),
                                 leading=9.5, wordWrap="CJK", fontName="Helvetica-Bold"),
    }

    def P(text, style="cell"):
        """Wrap a string in a Paragraph so the cell auto-wraps to column width.
        Escapes HTML-significant chars so '<' '>' '&' in asset names don't break."""
        s = "" if text is None else str(text)
        s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(s, styles[style])

    story: List[Any] = []
    # Cover
    story.append(Paragraph("Comparative Reports", styles["title"]))
    story.append(Paragraph("Railway Asset Inspection Management System",
                           styles["small"]))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        f"Generated for <b>{user.get('name', '—')}</b> ({user.get('role', '—').replace('_', ' ').title()})",
        styles["body"]))
    story.append(_filter_chips_para(bundle, styles))
    story.append(Spacer(1, 3 * mm))

    # ── Card A ──
    if req.sections.card_a:
        rows = _section_a_rows(bundle)
        story.append(Paragraph("A · MTTR by Asset Type", styles["h1"]))
        story.append(Paragraph(
            "Median/Mean repair hours per asset-type within your scope. "
            "Color: green = fast repair, red = slow.", styles["small"]))
        peer_max = max([r.get(req.stat) or 0 for r in rows] or [1])
        head = [P(h, "cell_b") for h in ["Rank", "Asset Type", "Median (hrs)", "Mean", "n", "Min", "Max", "Bar"]]
        data = [head]
        for i, r in enumerate(rows):
            v = r.get(req.stat)
            color_hex = "#10b981" if (v is not None and v / peer_max < 0.33) \
                else "#eab308" if (v is not None and v / peer_max < 0.66) \
                else "#f97316" if (v is not None and v / peer_max < 0.85) \
                else "#dc2626"
            d = _draw_horizontal_bar(v, peer_max, 35, 4, color_hex)
            data.append([P(str(i + 1)), P(r["label"]),
                         P(str(r.get("median") if r.get("median") is not None else "—")),
                         P(str(r.get("mean") if r.get("mean") is not None else "—")),
                         P(str(r.get("n") or 0)),
                         P(str(r.get("min") if r.get("min") is not None else "—")),
                         P(str(r.get("max") if r.get("max") is not None else "—")), d])
        tbl = Table(data, repeatRows=1, splitByRow=1,
                    colWidths=[12*mm, 50*mm, 22*mm, 18*mm, 12*mm, 14*mm, 14*mm, 38*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7c6b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (2, 1), (6, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)

    # ── Card B ──
    if req.sections.card_b:
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("B · Peer Supervisor Comparison", styles["h1"]))
        radar = await _radar_data(bundle)
        if not radar.get("series"):
            story.append(Paragraph("No peer-supervisor data in scope.", styles["small"]))
        else:
            if radar.get("anonymised"):
                story.append(Paragraph(
                    "Peers are anonymised — only your own polygon shows your name.",
                    styles["small"]))
            # Embed radar drawing
            drawing = _draw_radar(radar["axes"], radar["series"], size_mm=90)
            story.append(drawing)
            # Peer matrix table
            head = [P("Supervisor", "cell_b")] + [P(a["name"], "cell_b") for a in radar["axes"]]
            data = [head]
            for s in radar["series"]:
                row = [P(s["label"] + (" ★" if s.get("is_self") else ""))]
                row += [P(str(v.get("value")) if v.get("value") is not None else "—")
                        for v in s.get("values", [])]
                data.append(row)
            tbl = Table(data, repeatRows=1, splitByRow=1)
            ts = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7c6b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ]
            for ri, s in enumerate(radar["series"], start=1):
                if s.get("is_self"):
                    ts.append(("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#ccfbf1")))
            tbl.setStyle(TableStyle(ts))
            story.append(Spacer(1, 2 * mm))
            story.append(tbl)

    # ── Card C ──
    if req.sections.card_c_current or req.sections.card_c_full:
        story.append(PageBreak())
        story.append(Paragraph("C · Drilldown", styles["h1"]))
        if req.sections.card_c_current:
            story.append(Paragraph(
                f"Current view: level=<b>{req.drill_state.level}</b>", styles["small"]))
            cdata = await _drill_at(bundle, req.drill_state.level,
                                    req.drill_state.parent_id,
                                    req.drill_state.parent_asset_type_id)
            _render_drill_table(story, cdata, req.stat, styles)
        if req.sections.card_c_full:
            story.append(Spacer(1, 3 * mm))
            story.append(Paragraph("Full hierarchy (Station → Location → Asset Type)", styles["h2"]))
            await _render_full_hierarchy(story, bundle, styles)

    # ── Defective appendix ──
    defective_rows = _section_defective_rows(bundle) if req.sections.defective else []
    if req.sections.defective:
        story.append(PageBreak())
        story.append(Paragraph(
            f"Appendix · Defective Assets in Scope ({len(defective_rows)})",
            styles["h1"]))
        if not defective_rows:
            story.append(Paragraph("No open defects in scope.", styles["small"]))
        else:
            head = [P(h, "cell_b") for h in ["#", "Asset", "Type", "Station", "Location", "List", "Defective Since", "Days"]]
            data = [head]
            for i, r in enumerate(defective_rows, 1):
                ds = r["defective_since"][:16] if r["defective_since"] else "—"
                data.append([P(str(i)), P(r["asset_number"]), P(r["asset_type"]),
                             P(r["station"]), P(r["location"]),
                             P((r["list_type"] or "").upper() or "—"), P(ds),
                             P(f'{r["days_open"]:.1f}' if r["days_open"] is not None else "—")])
            tbl = Table(data, repeatRows=1, splitByRow=1,
                        colWidths=[8*mm, 32*mm, 28*mm, 22*mm, 32*mm, 12*mm, 26*mm, 12*mm])
            ts = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dc2626")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#fef2f2")]),
            ]
            for ri, r in enumerate(defective_rows, start=1):
                if r["list_type"] == "red":
                    ts.append(("TEXTCOLOR", (5, ri), (5, ri), colors.HexColor("#dc2626")))
                    ts.append(("FONTNAME", (5, ri), (5, ri), "Helvetica-Bold"))
                elif r["list_type"] == "orange":
                    ts.append(("TEXTCOLOR", (5, ri), (5, ri), colors.HexColor("#f97316")))
                    ts.append(("FONTNAME", (5, ri), (5, ri), "Helvetica-Bold"))
            tbl.setStyle(TableStyle(ts))
            story.append(tbl)

    # ── Last inspection appendix ──
    if req.sections.last_inspection:
        story.append(PageBreak())
        story.append(Paragraph("Appendix · Last Inspection per Asset", styles["h1"]))
        rows = await _last_inspection_rows(bundle, scope_asset_ids=None)
        if not rows:
            story.append(Paragraph("No inspections in scope.", styles["small"]))
        else:
            head = [P(h, "cell_b") for h in ["Asset", "Type", "Station", "Last Inspection", "Inspector", "Items", "Result"]]
            data = [head] + [
                [P(r["asset_number"]), P(r["asset_type"]), P(r["station"]),
                 P(r["last_at"]), P(r["inspector"]), P(str(r["item_count"])), P(r["result"])]
                for r in rows
            ]
            tbl = Table(data, repeatRows=1, splitByRow=1,
                        colWidths=[28*mm, 26*mm, 22*mm, 28*mm, 30*mm, 12*mm, 20*mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#f8fafc")]),
            ]))
            story.append(tbl)

    # ── Remarks appendix ──
    if req.sections.remarks:
        story.append(PageBreak())
        story.append(Paragraph("Appendix · Remarks (last 5 per defective asset)",
                               styles["h1"]))
        defective_ids = {r["asset_id"] for r in defective_rows} if defective_rows else \
            set(bundle["remarks_by_asset"].keys())
        any_rendered = False
        for aid in defective_ids:
            rems = bundle["remarks_by_asset"].get(aid, [])
            if not rems:
                continue
            a = bundle["asset_by_id"].get(aid) or {}
            any_rendered = True
            story.append(Paragraph(
                f"<b>{a.get('asset_number', '—')}</b> · "
                f"{(bundle['type_by_id'].get(a.get('asset_type_id')) or {}).get('name', '—')}",
                styles["h2"]))
            data = [[P(h, "cell_b") for h in ["When", "Role/Type", "Tag", "Body"]]]
            for r in rems:
                ts = str(r.get("created_at") or "")[:16]
                role_type = f"{(r.get('author_role') or '—')} · {(r.get('type') or '—')}"
                tag = r.get("tag") or "—"
                body = (r.get("body") or "")[:280]
                data.append([P(ts), P(role_type), P(tag), P(body)])
            tbl = Table(data, repeatRows=1, splitByRow=1,
                        colWidths=[24*mm, 32*mm, 28*mm, 96*mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#475569")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 2 * mm))
        if not any_rendered:
            story.append(Paragraph("No remarks for assets in current scope.",
                                   styles["small"]))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ─── Helpers — drill-fetch + radar + last inspections ────────────────────
async def _radar_data(bundle):
    """Fetch peer-supervisor radar data using same logic as the live endpoint."""
    user = bundle["user"]
    req: ExportRequest = bundle["req"]
    f_dt, t_dt = bundle["f_dt"], bundle["t_dt"]
    win = bundle["win"]
    type_by_id = bundle["type_by_id"]

    dept_id = req.dept_id or user.get("department_id")
    explicit_types = bundle["selected_type_ids"]
    axes = [{"id": tid, "name": (type_by_id.get(tid) or {}).get("name", "—")}
            for tid in explicit_types]
    sups = await users_collection.find(
        {"role": "supervisor", "department_id": dept_id, "is_active": True}).to_list(2000)
    if not sups:
        return {"axes": axes, "series": [], "anonymised": user.get("role") == "supervisor"}

    all_stns = set()
    for s in sups:
        all_stns |= set(s.get("assigned_stations") or [])
    asset_q = {"asset_type_id": {"$in": explicit_types}}
    if all_stns:
        asset_q["station_id"] = {"$in": list(all_stns)}
    assets = await assets_collection.find(asset_q).to_list(20000)
    asset_ids = [str(a["_id"]) for a in assets]
    asset_type_of = {str(a["_id"]): a.get("asset_type_id") for a in assets}
    ols = await orange_list_collection.find(
        {"asset_id": {"$in": asset_ids}, "status": "resolved"}).to_list(50000)

    is_anon = user.get("role") == "supervisor"
    series = []
    for i, sup in enumerate(sups):
        sid = str(sup["_id"])
        is_self = (sid == str(user["_id"]))
        vals = []
        for tid in explicit_types:
            sup_repairs = [ol for ol in ols
                           if ol.get("marked_working_by") == sid
                           and asset_type_of.get(ol.get("asset_id")) == tid]
            hours = _resolved_repair_hours(sup_repairs,
                                           {str(ol["asset_id"]) for ol in sup_repairs}, win)
            s = _hrs_stats(hours)
            vals.append({"asset_type_id": tid, "value": s.get(req.stat),
                         "n": s.get("n", 0)})
        series.append({
            "supervisor_id": sid if (not is_anon or is_self) else None,
            "label": sup.get("name") if (not is_anon or is_self) else f"Peer {i + 1}",
            "is_self": is_self, "values": vals,
        })
    return {"axes": axes, "series": series, "anonymised": is_anon}


async def _drill_at(bundle, level: str, parent_id: Optional[str],
                    parent_asset_type_id: Optional[str]):
    """Reuse comparative.grouped_drilldown logic — returns groups list."""
    from routers.comparative import grouped_drilldown
    user = bundle["user"]
    req: ExportRequest = bundle["req"]
    return await grouped_drilldown(
        user_id=str(user["_id"]),
        level=level,
        parent_id=parent_id,
        parent_asset_type_id=parent_asset_type_id,
        asset_type_ids=",".join(bundle["selected_type_ids"]) if bundle["selected_type_ids"] else None,
        dept_id=req.dept_id,
        window_days=req.window_days,
        stat=req.stat,
    )


def _render_drill_table(story, cdata, stat, styles):
    if not cdata or not cdata.get("groups"):
        story.append(Paragraph("No data at this drill level.", styles["small"]))
        return

    def _P(text, style="cell"):
        s = "" if text is None else str(text)
        s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(s, styles[style])

    bcrumb = " / ".join(b["label"] for b in cdata.get("breadcrumbs", []))
    story.append(Paragraph(f"<b>Path</b>: {bcrumb}", styles["small"]))
    is_station = cdata.get("level") == "station"
    head_label = "Station" if is_station else "Group"
    head = [_P(h, "cell_b") for h in [head_label, "Asset Type", f"{stat.title()} (hrs)", "n", "Min", "Max"]]
    data = [head]
    if is_station:
        for g in cdata["groups"]:
            for b in g["bars"]:
                if (b.get("n") or 0) == 0:
                    continue
                data.append([_P(g["label"]), _P(b.get("asset_type", "—")),
                             _P(b.get(stat)), _P(b.get("n")), _P(b.get("min")), _P(b.get("max"))])
    else:
        for g in cdata["groups"]:
            b = g["bars"][0] if g["bars"] else {}
            data.append([_P(g["label"]), _P(b.get("asset_type", "—")),
                         _P(b.get(stat)), _P(b.get("n")), _P(b.get("min")), _P(b.get("max"))])
    tbl = Table(data, repeatRows=1, splitByRow=1,
                colWidths=[55*mm, 40*mm, 24*mm, 14*mm, 18*mm, 18*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7c6b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tbl)


async def _render_full_hierarchy(story, bundle, styles):
    req: ExportRequest = bundle["req"]

    def _P(text, style="cell"):
        s = "" if text is None else str(text)
        s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(s, styles[style])

    station_data = await _drill_at(bundle, "station", None, None)
    for sg in (station_data.get("groups") or []):
        story.append(Paragraph(f"▸ {sg['label']}", styles["h2"]))
        loc_data = await _drill_at(bundle, "location_summary", sg["id"], None)
        if not loc_data.get("groups"):
            story.append(Paragraph("(no locations with data)", styles["muted"]))
            continue
        head = [_P(h, "cell_b") for h in ["Location", f"{req.stat.title()} (hrs)", "n", "Min", "Max"]]
        data = [head] + [
            [_P(g["label"]), _P(g["bars"][0].get(req.stat)), _P(g["bars"][0].get("n")),
             _P(g["bars"][0].get("min")), _P(g["bars"][0].get("max"))]
            for g in loc_data["groups"]
        ]
        tbl = Table(data, repeatRows=1, colWidths=[60*mm, 22*mm, 12*mm, 14*mm, 14*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#475569")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 2 * mm))


async def _last_inspection_rows(bundle, scope_asset_ids=None):
    out = []
    inspector_ids = list({ins.get("inspector_id") for ins in bundle["last_inspections"].values()
                          if ins.get("inspector_id")})
    inspectors = {}
    if inspector_ids:
        users = await users_collection.find(
            {"_id": {"$in": [ObjectId(i) for i in inspector_ids]}}).to_list(2000)
        inspectors = {str(u["_id"]): u.get("name", "—") for u in users}
    for aid, ins in bundle["last_inspections"].items():
        if scope_asset_ids is not None and aid not in scope_asset_ids:
            continue
        a = bundle["asset_by_id"].get(aid) or {}
        st = bundle["station_by_id"].get(a.get("station_id")) or {}
        tp = bundle["type_by_id"].get(a.get("asset_type_id")) or {}
        items = ins.get("items") or []
        result_codes = [str(it.get("status", "")).lower() for it in items]
        if any(s == "not_ok" for s in result_codes):
            result = "NOT OK"
        elif any(s == "needs_repair" for s in result_codes):
            result = "NEEDS REPAIR"
        else:
            result = "OK" if items else "—"
        out.append({
            "asset_number": a.get("asset_number", "—"),
            "asset_type": tp.get("name", "—"),
            "station": st.get("name", "—"),
            "last_at": (str(ins.get("inspection_at") or "")[:16]),
            "inspector": inspectors.get(str(ins.get("inspector_id")), "—"),
            "item_count": len(items),
            "result": result,
        })
    out.sort(key=lambda r: r["last_at"], reverse=True)
    return out


# ─── Excel ────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="0E7C6B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
STRIPE_FILL = PatternFill("solid", fgColor="F8FAFC")  # subtle slate-50 for odd rows
THIN_BORDER = Border(left=Side(style="thin", color="CBD5E1"),
                     right=Side(style="thin", color="CBD5E1"),
                     top=Side(style="thin", color="CBD5E1"),
                     bottom=Side(style="thin", color="CBD5E1"))


def _style_header(ws, row=1):
    for c in ws[row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER


def _stripe_rows(ws, start_row: int, end_row: int = None,
                 skip_rows: set = None):
    """Apply subtle alternating fill to data rows. `skip_rows` are 1-indexed
    row numbers that already have a highlight (e.g. self-row in peer matrix)
    and should not be overwritten."""
    if end_row is None:
        end_row = ws.max_row
    skip = skip_rows or set()
    for r in range(start_row, end_row + 1):
        if r in skip:
            continue
        # Stripe every OTHER data row: first data row stays white, second
        # gets the stripe, etc.
        if (r - start_row) % 2 == 1:
            for c in ws[r]:
                if c.fill.fgColor and c.fill.fgColor.rgb in ("00000000", None):
                    c.fill = STRIPE_FILL


def _freeze_below_header(ws, header_row: int = 1):
    """Freeze panes so the header row stays visible while scrolling."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def _autofit(ws, max_w=60):
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_w, max(8, length + 2))


async def _build_excel(user_id: str, req: ExportRequest) -> bytes:
    bundle = await _build_data_bundle(user_id, req)
    user = bundle["user"]
    wb = Workbook()
    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Comparative Reports — Export"])
    ws["A1"].font = Font(bold=True, size=14, color="0E7C6B")
    ws.append([])
    ws.append(["User", user.get("name", "—")])
    ws.append(["Role", user.get("role", "—").replace("_", " ").title()])
    ws.append(["Window", req.window_days])
    ws.append(["Stat", req.stat])
    if req.dept_id:
        d = bundle["dept_by_id"].get(req.dept_id) or {}
        ws.append(["Department", d.get("name", "—")])
    types_str = ", ".join((bundle["type_by_id"].get(t) or {}).get("name", "—")
                          for t in bundle["selected_type_ids"])
    ws.append(["Asset Types", types_str])
    ws.append(["Generated", now_ist().strftime("%d %b %Y, %H:%M IST")])
    _autofit(ws)

    # Card A
    if req.sections.card_a:
        ws = wb.create_sheet("By Asset Type")
        ws.append(["Rank", "Asset Type", "Median (hrs)", "Mean", "n", "Min", "Max", "p75", "p90"])
        _style_header(ws)
        for i, r in enumerate(_section_a_rows(bundle), 1):
            ws.append([i, r["label"], r.get("median"), r.get("mean"),
                       r.get("n"), r.get("min"), r.get("max"),
                       r.get("p75"), r.get("p90")])
        _stripe_rows(ws, start_row=2)
        _freeze_below_header(ws)
        _autofit(ws)

    # Card B
    if req.sections.card_b:
        radar = await _radar_data(bundle)
        ws = wb.create_sheet("Peer Matrix")
        head = ["Supervisor"] + [a["name"] for a in radar.get("axes", [])]
        ws.append(head)
        _style_header(ws)
        self_rows = set()
        for s in radar.get("series", []):
            row = [s["label"] + (" ★" if s.get("is_self") else "")]
            row += [(v.get("value") if v.get("value") is not None else None)
                    for v in s.get("values", [])]
            ws.append(row)
            if s.get("is_self"):
                self_rows.add(ws.max_row)
                for c in ws[ws.max_row]:
                    c.fill = PatternFill("solid", fgColor="CCFBF1")
        _stripe_rows(ws, start_row=2, skip_rows=self_rows)
        _freeze_below_header(ws)
        if radar.get("anonymised"):
            ws.append([])
            ws.append(["Note: peers anonymised (SUP role)."])
        _autofit(ws)

    # Card C — full hierarchy if requested, else current drill
    if req.sections.card_c_full:
        ws = wb.create_sheet("Drilldown — Full")
        ws.append(["Station", "Location", "Asset Type", "Asset Number",
                   f"{req.stat.title()} (hrs)", "n", "Min", "Max"])
        _style_header(ws)
        # Iterate per asset for completeness
        for a in bundle["asset_by_id"].values():
            aid = str(a["_id"])
            hours = _resolved_repair_hours(bundle["ols_resolved"], aid, bundle["win"])
            s = _hrs_stats(hours)
            st = bundle["station_by_id"].get(a.get("station_id")) or {}
            loc = bundle["location_by_id"].get(a.get("location_id")) or {}
            tp = bundle["type_by_id"].get(a.get("asset_type_id")) or {}
            ws.append([st.get("name", "—"), loc.get("name", "—"),
                       tp.get("name", "—"), a.get("asset_number"),
                       s.get(req.stat), s.get("n"), s.get("min"), s.get("max")])
        _stripe_rows(ws, start_row=2)
        _freeze_below_header(ws)
        _autofit(ws)
    elif req.sections.card_c_current:
        cdata = await _drill_at(bundle, req.drill_state.level,
                                req.drill_state.parent_id,
                                req.drill_state.parent_asset_type_id)
        ws = wb.create_sheet("Drilldown")
        bcrumb = " / ".join(b["label"] for b in cdata.get("breadcrumbs", []))
        ws.append([f"Drill Path: {bcrumb}"])
        ws.append([])
        ws.append(["Group", "Asset Type", f"{req.stat.title()} (hrs)", "n", "Min", "Max"])
        _style_header(ws, row=3)
        for g in cdata.get("groups", []):
            for b in g["bars"]:
                ws.append([g["label"], b.get("asset_type", "—"),
                           b.get(req.stat), b.get("n"), b.get("min"), b.get("max")])
        _stripe_rows(ws, start_row=4)
        _freeze_below_header(ws, header_row=3)
        _autofit(ws)

    # Defective
    defective = _section_defective_rows(bundle) if req.sections.defective else []
    if req.sections.defective:
        ws = wb.create_sheet("Defective Only")
        ws.append(["Asset", "Type", "Station", "Location", "List Type",
                   "Status", "Defective Since", "Days Open"])
        _style_header(ws)
        highlighted_rows = set()
        for r in defective:
            ws.append([r["asset_number"], r["asset_type"], r["station"],
                       r["location"], r["list_type"].upper(), r["status"],
                       r["defective_since"][:16] if r["defective_since"] else "",
                       r["days_open"]])
            row = ws[ws.max_row]
            if r["list_type"] == "red":
                highlighted_rows.add(ws.max_row)
                for c in row:
                    c.fill = PatternFill("solid", fgColor="FECACA")
            elif r["list_type"] == "orange":
                highlighted_rows.add(ws.max_row)
                for c in row:
                    c.fill = PatternFill("solid", fgColor="FED7AA")
        _stripe_rows(ws, start_row=2, skip_rows=highlighted_rows)
        _freeze_below_header(ws)
        _autofit(ws)

    # Last inspection
    if req.sections.last_inspection:
        ws = wb.create_sheet("Last Inspections")
        ws.append(["Asset", "Type", "Station", "Last Inspection", "Inspector",
                   "Items", "Result"])
        _style_header(ws)
        rows = await _last_inspection_rows(bundle)
        for r in rows:
            ws.append([r["asset_number"], r["asset_type"], r["station"],
                       r["last_at"], r["inspector"], r["item_count"], r["result"]])
        _stripe_rows(ws, start_row=2)
        _freeze_below_header(ws)
        _autofit(ws)

    # Remarks
    if req.sections.remarks:
        ws = wb.create_sheet("Remarks")
        ws.append(["Asset", "When", "Author Role", "Type", "Tag", "Body"])
        _style_header(ws)
        defective_ids = {r["asset_id"] for r in defective} if defective else \
            set(bundle["remarks_by_asset"].keys())
        for aid in defective_ids:
            rems = bundle["remarks_by_asset"].get(aid, [])
            a = bundle["asset_by_id"].get(aid) or {}
            for r in rems:
                ws.append([
                    a.get("asset_number", "—"),
                    str(r.get("created_at") or "")[:16],
                    r.get("author_role", "—"),
                    r.get("type", "—"),
                    r.get("tag", "—"),
                    (r.get("body") or "")[:300],
                ])
        _stripe_rows(ws, start_row=2)
        _freeze_below_header(ws)
        _autofit(ws, max_w=80)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ─── Endpoints ────────────────────────────────────────────────────────────
@router.post("/api/reports/comparative/export/pdf/{user_id}")
async def export_pdf(user_id: str, req: ExportRequest):
    pdf_bytes = await _build_pdf(user_id, req)
    fname = f"comparative_{user_id[:8]}_{now_ist().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/api/reports/comparative/export/excel/{user_id}")
async def export_excel(user_id: str, req: ExportRequest):
    xlsx_bytes = await _build_excel(user_id, req)
    fname = f"comparative_{user_id[:8]}_{now_ist().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
