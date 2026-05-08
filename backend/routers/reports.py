"""
Reports module — role-scoped health aggregation + PDF/Excel exports.

Endpoints:
  GET  /api/reports/health/{user_id}      — aggregated tree based on role
  GET  /api/reports/export/pdf/{user_id}  — PDF download (ReportLab)
  GET  /api/reports/export/excel/{user_id}— XLSX download (openpyxl)

Role-scoped tree shape:
  • SUP    → stations: [ {station, dept_pct, asset_type_rings[], locations[]} ]
  • RO     → supervisors: [ {sup_summary, drillable=True} ]
  • ASUP   → supervisors: [ {sup_summary} ]   (cross-dept umbrella, station-scoped)
  • Admin/SA → ros: [ {ro_summary, dept_rings[], supervisor_bars[]} ]
"""
import io
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Any

from bson import ObjectId
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from database import (
    users_collection, assets_collection, asset_types_collection,
    stations_collection, locations_collection, departments_collection,
    orange_list_collection, now_ist,
)

router = APIRouter()

# ─── Color gradient (mirrors frontend HealthGradient.js) ────────────────────
def health_color(pct: float) -> str:
    """Return hex color for a working % on the gradient."""
    if pct >= 100:
        return "#10b981"      # pure green
    if pct >= 95:
        # 95→100 yellow→green
        t = (pct - 95) / 5.0
        return _lerp("#eab308", "#4ade80", t)
    if pct >= 90:
        t = (pct - 90) / 5.0
        return _lerp("#f97316", "#eab308", t)
    if pct >= 80:
        t = (pct - 80) / 10.0
        return _lerp("#dc2626", "#f97316", t)
    return "#7f1d1d"           # deep red below 80


def _lerp(a: str, b: str, t: float) -> str:
    t = max(0.0, min(1.0, t))
    ar, ag, ab = int(a[1:3], 16), int(a[3:5], 16), int(a[5:7], 16)
    br, bg, bb = int(b[1:3], 16), int(b[3:5], 16), int(b[5:7], 16)
    r, g, bl = int(ar + (br - ar) * t), int(ag + (bg - ag) * t), int(ab + (bb - ab) * t)
    return f"#{r:02x}{g:02x}{bl:02x}"


# ─── Health classification helper ───────────────────────────────────────────
def _classify(asset, ol_open):
    """Return one of: 'working', 'yellow', 'orange', 'red'."""
    st = asset.get("status")
    if st == "working":
        return "working"
    if st == "pending_approval":
        return "yellow"
    ds = ol_open.get("defective_since") if ol_open else asset.get("defective_since")
    if not ds:
        return "orange"
    if isinstance(ds, str):
        try:
            ds = datetime.fromisoformat(ds.replace("Z", "").replace("+00:00", ""))
        except Exception:
            return "orange"
    if ds.tzinfo is not None:
        ds = ds.replace(tzinfo=None)
    hours = (now_ist() - ds).total_seconds() / 3600
    return "red" if hours > 24 else "orange"


def _empty_bucket():
    return {"working": 0, "yellow": 0, "orange": 0, "red": 0}


def _parse_dt(v):
    """Parse a stored datetime value (str or datetime) into a naive datetime."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v.replace("Z", "").replace("+00:00", ""))
        except Exception:
            return None
    return None


def _compute_30day_trend(station_assets: List[dict], all_ols_by_asset: Dict[str, list],
                        now_dt: datetime) -> List[float]:
    """Compute per-day % working (incl. yellow) for the last 30 days.

    Returns a list of 30 floats — index 0 = 29 days ago, index 29 = today.
    An asset is "defective" at time T if any OL entry has:
        defective_since <= T  AND  end > T   (end = earliest of marked_working_at, approved_at; or open)
    Yellow & Working both count as "working" in the percentage.
    """
    total = len(station_assets)
    if total == 0:
        return [100.0] * 30

    # Pre-build intervals per asset: list of (defective_since, end_or_None)
    intervals_by_asset: Dict[str, list] = {}
    for a in station_assets:
        aid = str(a["_id"])
        ivs = []
        for ol in all_ols_by_asset.get(aid, []):
            ds = _parse_dt(ol.get("defective_since"))
            if not ds:
                continue
            mw = _parse_dt(ol.get("marked_working_at"))
            ap = _parse_dt(ol.get("approved_at"))
            end = None
            for cand in (mw, ap):
                if cand and (end is None or cand < end):
                    end = cand
            ivs.append((ds, end))
        intervals_by_asset[aid] = ivs

    eod_today = now_dt.replace(hour=23, minute=59, second=59, microsecond=0)
    trend: List[float] = []
    for offset in range(29, -1, -1):
        T = eod_today - timedelta(days=offset)
        defective = 0
        for a in station_assets:
            aid = str(a["_id"])
            for ds, end in intervals_by_asset.get(aid, []):
                if ds <= T and (end is None or end > T):
                    defective += 1
                    break
        working = total - defective
        trend.append(round((working / total * 100), 1))
    return trend


def _bucket_pct(b: Dict[str, int]) -> Dict[str, Any]:
    """Center-percent: (working + yellow) / total. Add color and total."""
    total = sum(b.values())
    pct = round(((b["working"] + b["yellow"]) / total * 100), 1) if total else 100.0
    return {
        **b,
        "total": total,
        "pct_working": pct,
        "color": health_color(pct),
    }


# ─── Core builder: classify an asset & accumulate buckets ────────────────────
async def _load_universe():
    """Load everything once. Small dataset → single fetch is cheap."""
    assets = await assets_collection.find({}).to_list(10000)
    types = await asset_types_collection.find({}).to_list(2000)
    stations = await stations_collection.find({}).to_list(2000)
    locations = await locations_collection.find({}).to_list(2000)
    depts = await departments_collection.find({}).to_list(500)
    open_ols = await orange_list_collection.find({"status": {"$ne": "resolved"}}).to_list(20000)
    # ALL OLs (incl. resolved) — needed for 30-day trend reconstruction
    all_ols = await orange_list_collection.find({}).to_list(50000)
    users = await users_collection.find({"role": {"$in": ["supervisor", "reporting_officer", "approving_supervisor"]}}).to_list(2000)
    type_by_id = {str(t["_id"]): t for t in types}
    station_by_id = {str(s["_id"]): s for s in stations}
    location_by_id = {str(l["_id"]): l for l in locations}
    dept_by_id = {str(d["_id"]): d for d in depts}
    user_by_id = {str(u["_id"]): u for u in users}
    ol_by_asset = {ol["asset_id"]: ol for ol in open_ols}
    # Group ALL ols by asset_id (list, since multiple resolved entries can exist)
    all_ols_by_asset: Dict[str, list] = defaultdict(list)
    for ol in all_ols:
        if ol.get("asset_id"):
            all_ols_by_asset[ol["asset_id"]].append(ol)
    return {
        "assets": assets, "types": types, "stations": stations, "locations": locations,
        "type_by_id": type_by_id, "station_by_id": station_by_id,
        "location_by_id": location_by_id, "dept_by_id": dept_by_id,
        "user_by_id": user_by_id, "ol_by_asset": ol_by_asset,
        "all_ols_by_asset": all_ols_by_asset,
    }


def _filter_assets_for_user(U: dict, user: dict) -> List[dict]:
    """Apply role-scoping to the assets list."""
    role = user.get("role")
    assigned = set(user.get("assigned_stations") or [])
    dept_id = user.get("department_id")
    if role in ("admin", "superadmin"):
        return U["assets"]
    if role == "approving_supervisor":
        return [a for a in U["assets"] if a.get("station_id") in assigned]
    if role == "reporting_officer":
        # Same dept + assigned stations
        return [
            a for a in U["assets"]
            if a.get("station_id") in assigned
            and U["type_by_id"].get(a.get("asset_type_id"), {}).get("department_id") == dept_id
        ]
    if role == "supervisor":
        return [
            a for a in U["assets"]
            if a.get("station_id") in assigned
            and U["type_by_id"].get(a.get("asset_type_id"), {}).get("department_id") == dept_id
        ]
    return []


def _build_station_card(U: dict, station_id: str, scoped_assets: List[dict],
                        ring_grouping: str = "asset_type") -> Optional[dict]:
    """
    Build the card structure for a single station.
    `ring_grouping` = 'asset_type' (SUP/RO) or 'department' (Admin/ASUP).
    """
    station_assets = [a for a in scoped_assets if a.get("station_id") == station_id]
    if not station_assets:
        return None  # G13: hide empty stations

    summary = _empty_bucket()
    rings: Dict[str, Dict[str, int]] = defaultdict(_empty_bucket)
    locations: Dict[str, Dict[str, int]] = defaultdict(_empty_bucket)

    for a in station_assets:
        cls = _classify(a, U["ol_by_asset"].get(str(a["_id"])))
        summary[cls] += 1

        # Ring grouping
        if ring_grouping == "department":
            t = U["type_by_id"].get(a.get("asset_type_id"), {})
            d = U["dept_by_id"].get(t.get("department_id"), {})
            key = d.get("name", "—")
        else:  # asset_type
            t = U["type_by_id"].get(a.get("asset_type_id"), {})
            key = t.get("name", "—")
        rings[key][cls] += 1

        # Locations
        loc = U["location_by_id"].get(a.get("location_id"), {})
        loc_name = loc.get("name", "—")
        locations[loc_name][cls] += 1

    summary_b = _bucket_pct(summary)

    # Ring list — keep stable order: most defective first
    ring_list = []
    for name, b in rings.items():
        bp = _bucket_pct(b)
        defect_count = bp["yellow"] + bp["orange"] + bp["red"]
        ring_list.append({"name": name, **bp, "_defect_count": defect_count})
    ring_list.sort(key=lambda r: -r["_defect_count"])
    for r in ring_list:
        r.pop("_defect_count", None)

    # Location list — worst-first (D8)
    loc_list = []
    for name, b in locations.items():
        bp = _bucket_pct(b)
        defect_count = bp["yellow"] + bp["orange"] + bp["red"]
        loc_list.append({"name": name, **bp, "_defect_count": defect_count})
    loc_list.sort(key=lambda l: (-l["_defect_count"], -l["total"]))
    for l in loc_list:
        l.pop("_defect_count", None)

    station = U["station_by_id"].get(station_id, {})
    trend_30d = _compute_30day_trend(station_assets, U["all_ols_by_asset"], now_ist())
    return {
        "station_id": station_id,
        "station_name": station.get("name", "—"),
        "summary": summary_b,
        "rings": ring_list,
        "locations": loc_list,
        "trend_30d": trend_30d,
    }


def _build_supervisor_summary(U: dict, sup: dict) -> dict:
    """Compact mini-card data for one supervisor."""
    assets = _filter_assets_for_user(U, sup)
    bucket = _empty_bucket()
    station_buckets: Dict[str, Dict[str, int]] = defaultdict(_empty_bucket)
    for a in assets:
        cls = _classify(a, U["ol_by_asset"].get(str(a["_id"])))
        bucket[cls] += 1
        station_buckets[a.get("station_id")][cls] += 1
    summary = _bucket_pct(bucket)
    stations = []
    for sid, b in station_buckets.items():
        s = U["station_by_id"].get(sid, {})
        bp = _bucket_pct(b)
        stations.append({"station_id": sid, "station_name": s.get("name", "—"), **bp})
    stations.sort(key=lambda s: (-(s["yellow"] + s["orange"] + s["red"]), -s["total"]))
    return {
        "user_id": str(sup["_id"]),
        "name": sup.get("name", "—"),
        "employee_id": sup.get("employee_id", ""),
        "role": sup.get("role"),
        "department_id": sup.get("department_id"),
        "station_count": len(station_buckets),
        "summary": summary,
        "stations": stations,  # used for "supervisor_bars" / drill-down
    }


# ════════════════════════════════════════════════════════════════════════════
@router.get("/api/reports/health/{user_id}")
async def reports_health(user_id: str, drill_user_id: Optional[str] = Query(None)):
    """Role-scoped health tree.

    Optional `drill_user_id` param: when an Admin/SA/RO views a SUP's drilled-down
    cards, pass the SUP's user_id here.
    """
    try:
        user = await users_collection.find_one({"_id": ObjectId(user_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user_id")
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    U = await _load_universe()
    role = user.get("role")
    generated_at = now_ist()

    # ── Drill-down: render SUP-style station cards for the target user ──────
    if drill_user_id:
        try:
            target = await users_collection.find_one({"_id": ObjectId(drill_user_id)})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid drill_user_id")
        if not target:
            raise HTTPException(status_code=404, detail="Drill target not found")
        scoped = _filter_assets_for_user(U, target)
        station_ids = sorted({a.get("station_id") for a in scoped})
        ring_grouping = "department" if target.get("role") in ("admin", "superadmin", "approving_supervisor") else "asset_type"
        cards = [c for sid in station_ids if (c := _build_station_card(U, sid, scoped, ring_grouping))]
        return {
            "view": "stations",
            "viewer": {"user_id": user_id, "role": role},
            "target": {
                "user_id": str(target["_id"]),
                "name": target.get("name"),
                "employee_id": target.get("employee_id"),
                "role": target.get("role"),
            },
            "stations": cards,
            "generated_at": generated_at.isoformat(),
        }

    # ── Standard role views ──────────────────────────────────────────────────
    if role == "supervisor":
        scoped = _filter_assets_for_user(U, user)
        station_ids = sorted({a.get("station_id") for a in scoped})
        cards = [c for sid in station_ids if (c := _build_station_card(U, sid, scoped, "asset_type"))]
        return {"view": "stations", "viewer": {"user_id": user_id, "role": role},
                "stations": cards, "generated_at": generated_at.isoformat()}

    if role in ("reporting_officer", "approving_supervisor"):
        # Supervisors who report to this RO (or for ASUP, all supervisors at their stations)
        if role == "reporting_officer":
            sups = [u for u in U["user_by_id"].values()
                    if u.get("role") == "supervisor" and u.get("reports_to_id") == user_id]
        else:  # ASUP — all SUPs at their assigned stations (cross-dept umbrella)
            assigned = set(user.get("assigned_stations") or [])
            sups = [u for u in U["user_by_id"].values()
                    if u.get("role") == "supervisor" and (set(u.get("assigned_stations") or []) & assigned)]
        sup_summaries = [_build_supervisor_summary(U, s) for s in sups]
        sup_summaries.sort(key=lambda s: -(s["summary"]["yellow"] + s["summary"]["orange"] + s["summary"]["red"]))
        return {"view": "supervisors", "viewer": {"user_id": user_id, "role": role},
                "supervisors": sup_summaries, "generated_at": generated_at.isoformat()}

    if role in ("admin", "superadmin"):
        # Group by RO; each RO card = department-wise rings + supervisor bars
        ros = [u for u in U["user_by_id"].values() if u.get("role") == "reporting_officer"]
        cards = []
        for ro in ros:
            sups = [u for u in U["user_by_id"].values()
                    if u.get("role") == "supervisor" and u.get("reports_to_id") == str(ro["_id"])]
            ro_assets: List[dict] = []
            for s in sups:
                ro_assets.extend(_filter_assets_for_user(U, s))
            # de-dupe by asset _id
            seen, deduped = set(), []
            for a in ro_assets:
                aid = str(a["_id"])
                if aid not in seen:
                    seen.add(aid)
                    deduped.append(a)
            if not deduped:
                continue
            # Rings = department-wise (C4 for Admin/ASUP)
            summary = _empty_bucket()
            dept_rings: Dict[str, Dict[str, int]] = defaultdict(_empty_bucket)
            sup_bars: Dict[str, Dict[str, int]] = defaultdict(_empty_bucket)
            sup_meta: Dict[str, dict] = {}
            for a in deduped:
                cls = _classify(a, U["ol_by_asset"].get(str(a["_id"])))
                summary[cls] += 1
                t = U["type_by_id"].get(a.get("asset_type_id"), {})
                d = U["dept_by_id"].get(t.get("department_id"), {})
                dept_rings[d.get("name", "—")][cls] += 1
                # Find which supervisor of this RO covers this asset
                for s in sups:
                    if a.get("station_id") in (s.get("assigned_stations") or []) \
                       and t.get("department_id") == s.get("department_id"):
                        sid = str(s["_id"])
                        sup_bars[sid][cls] += 1
                        sup_meta[sid] = {"name": s.get("name"), "employee_id": s.get("employee_id")}
                        break

            # Format
            ring_list = sorted(
                [{"name": k, **_bucket_pct(v)} for k, v in dept_rings.items()],
                key=lambda r: -(r["yellow"] + r["orange"] + r["red"])
            )
            sup_bar_list = []
            for sid, b in sup_bars.items():
                bp = _bucket_pct(b)
                m = sup_meta.get(sid, {})
                sup_bar_list.append({"user_id": sid, **m, **bp})
            sup_bar_list.sort(key=lambda s: -(s["yellow"] + s["orange"] + s["red"]))

            cards.append({
                "user_id": str(ro["_id"]),
                "name": ro.get("name"),
                "employee_id": ro.get("employee_id"),
                "role": "reporting_officer",
                "department_id": ro.get("department_id"),
                "supervisor_count": len(sups),
                "station_count": len({a.get("station_id") for a in deduped}),
                "summary": _bucket_pct(summary),
                "rings": ring_list,
                "supervisor_bars": sup_bar_list,
            })
        cards.sort(key=lambda c: -(c["summary"]["yellow"] + c["summary"]["orange"] + c["summary"]["red"]))
        return {"view": "ros", "viewer": {"user_id": user_id, "role": role},
                "ros": cards, "generated_at": generated_at.isoformat()}

    raise HTTPException(status_code=403, detail=f"No report view for role '{role}'")


# ════════════════════════════════════════════════════════════════════════════
# EXPORTS — PDF & EXCEL
# ════════════════════════════════════════════════════════════════════════════

def _pdf_color(hex_str: str):
    from reportlab.lib.colors import HexColor
    return HexColor(hex_str)


@router.get("/api/reports/export/pdf/{user_id}")
async def export_pdf(user_id: str, drill_user_id: Optional[str] = Query(None)):
    """Server-side PDF (ReportLab). Cover summary + per-card pages."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    data = await reports_health(user_id, drill_user_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm,
                            topMargin=18*mm, bottomMargin=18*mm,
                            title="Asset Health Report")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], textColor=_pdf_color("#0e7c6b"), fontSize=18)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, textColor=_pdf_color("#0f172a"))
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, textColor=_pdf_color("#64748b"))
    normal = styles["Normal"]

    elements: List[Any] = []
    elements.append(Paragraph("Asset Health Report", title_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f"Viewer: {data['viewer']['role'].upper()} · Generated: {data['generated_at']}",
        small,
    ))
    elements.append(Spacer(1, 12))

    def _summary_table(b):
        t = Table(
            [["Total", "Working", "Yellow", "Orange", "Red", "% Working"],
             [b["total"], b["working"], b["yellow"], b["orange"], b["red"], f"{b['pct_working']:.1f}%"]],
            colWidths=[24*mm]*6,
        )
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _pdf_color("#0e7c6b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), _pdf_color("#ffffff")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONT", (0, 1), (-1, 1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, _pdf_color("#cbd5e1")),
            ("BACKGROUND", (5, 1), (5, 1), _pdf_color(b["color"])),
            ("TEXTCOLOR", (5, 1), (5, 1), _pdf_color("#ffffff")),
        ]))
        return t

    def _ring_table(rings, kind="Asset type"):
        if not rings:
            return Paragraph("No items", small)
        rows = [[kind, "Total", "W", "Y", "O", "R", "% Working"]]
        for r in rings:
            rows.append([r["name"], r["total"], r["working"], r["yellow"], r["orange"], r["red"], f"{r['pct_working']:.1f}%"])
        t = Table(rows, colWidths=[40*mm, 20*mm, 18*mm, 18*mm, 18*mm, 18*mm, 22*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _pdf_color("#f1f5f9")),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.3, _pdf_color("#cbd5e1")),
        ]))
        return t

    def _render_station_card(card):
        elements.append(Paragraph(f"<b>{card['station_name']}</b>", h2))
        elements.append(_summary_table(card["summary"]))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("<b>Asset-type / Department breakdown</b>", normal))
        elements.append(_ring_table(card["rings"]))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("<b>Location-wise health</b> (worst first)", normal))
        elements.append(_ring_table(card["locations"], kind="Location"))
        elements.append(Spacer(1, 14))

    view = data["view"]
    if view == "stations":
        elements.append(Paragraph("Per-station health cards", h2))
        elements.append(Spacer(1, 6))
        for card in data["stations"]:
            _render_station_card(card)
            elements.append(PageBreak())
    elif view == "supervisors":
        elements.append(Paragraph("Per-supervisor summary", h2))
        elements.append(Spacer(1, 6))
        rows = [["Employee", "Name", "Stations", "Total", "W", "Y", "O", "R", "% Working"]]
        for s in data["supervisors"]:
            sm = s["summary"]
            rows.append([s["employee_id"], s["name"], s["station_count"], sm["total"],
                         sm["working"], sm["yellow"], sm["orange"], sm["red"], f"{sm['pct_working']:.1f}%"])
        elements.append(Table(rows, colWidths=[22*mm]*9, style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _pdf_color("#0e7c6b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), _pdf_color("#ffffff")),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, _pdf_color("#cbd5e1")),
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ])))
    elif view == "ros":
        elements.append(Paragraph("Per-RO summary", h2))
        elements.append(Spacer(1, 6))
        for ro in data["ros"]:
            elements.append(Paragraph(f"<b>{ro['name']}</b> ({ro['employee_id']})  ·  {ro['supervisor_count']} supervisors  ·  {ro['station_count']} stations", h2))
            elements.append(_summary_table(ro["summary"]))
            elements.append(Spacer(1, 6))
            elements.append(Paragraph("<b>Department breakdown</b>", normal))
            elements.append(_ring_table(ro["rings"], kind="Department"))
            elements.append(Spacer(1, 6))
            if ro["supervisor_bars"]:
                rows = [["Supervisor", "Total", "W", "Y", "O", "R", "% Working"]]
                for s in ro["supervisor_bars"]:
                    rows.append([f"{s.get('name','-')} ({s.get('employee_id','-')})",
                                 s["total"], s["working"], s["yellow"], s["orange"], s["red"],
                                 f"{s['pct_working']:.1f}%"])
                elements.append(Paragraph("<b>Supervisor-wise health</b>", normal))
                elements.append(Table(rows, colWidths=[55*mm, 16*mm, 14*mm, 14*mm, 14*mm, 14*mm, 22*mm], style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), _pdf_color("#f1f5f9")),
                    ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.3, _pdf_color("#cbd5e1")),
                ])))
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    fname = f"asset-health-report-{now_ist().strftime('%Y%m%d-%H%M')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/api/reports/export/excel/{user_id}")
async def export_excel(user_id: str, drill_user_id: Optional[str] = Query(None)):
    """Multi-sheet xlsx: Summary, Departments, Stations, Locations, Assets (flat)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    data = await reports_health(user_id, drill_user_id)
    U = await _load_universe()

    # Pull viewer for asset filtering
    user = await users_collection.find_one({"_id": ObjectId(user_id)})
    target_user = user
    if drill_user_id:
        target_user = await users_collection.find_one({"_id": ObjectId(drill_user_id)}) or user
    asset_pool = _filter_assets_for_user(U, target_user)

    wb = Workbook()
    teal_fill = PatternFill("solid", fgColor="0E7C6B")
    bold_white = Font(bold=True, color="FFFFFF")

    def _hdr(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = teal_fill
            c.font = bold_white
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # Summary
    ws = wb.active
    ws.title = "Summary"
    _hdr(ws, ["Metric", "Value"])
    rows = [
        ["Generated at", data["generated_at"]],
        ["Viewer role", data["viewer"]["role"]],
        ["View", data["view"]],
    ]
    for r in rows:
        ws.append(r)

    # View-specific sheets
    if data["view"] == "stations":
        ws2 = wb.create_sheet("Stations")
        _hdr(ws2, ["Station", "Total", "Working", "Yellow", "Orange", "Red", "% Working"])
        for s in data["stations"]:
            sm = s["summary"]
            ws2.append([s["station_name"], sm["total"], sm["working"], sm["yellow"], sm["orange"], sm["red"], sm["pct_working"]])

        ws3 = wb.create_sheet("Locations")
        _hdr(ws3, ["Station", "Location", "Total", "Working", "Yellow", "Orange", "Red", "% Working"])
        for s in data["stations"]:
            for l in s["locations"]:
                ws3.append([s["station_name"], l["name"], l["total"], l["working"], l["yellow"], l["orange"], l["red"], l["pct_working"]])

        ws4 = wb.create_sheet("Asset Types")
        _hdr(ws4, ["Station", "Asset Type", "Total", "Working", "Yellow", "Orange", "Red", "% Working"])
        for s in data["stations"]:
            for r in s["rings"]:
                ws4.append([s["station_name"], r["name"], r["total"], r["working"], r["yellow"], r["orange"], r["red"], r["pct_working"]])

    elif data["view"] == "supervisors":
        ws2 = wb.create_sheet("Supervisors")
        _hdr(ws2, ["Employee", "Name", "Stations", "Total", "W", "Y", "O", "R", "% Working"])
        for s in data["supervisors"]:
            sm = s["summary"]
            ws2.append([s["employee_id"], s["name"], s["station_count"], sm["total"],
                        sm["working"], sm["yellow"], sm["orange"], sm["red"], sm["pct_working"]])

    elif data["view"] == "ros":
        ws2 = wb.create_sheet("ROs")
        _hdr(ws2, ["Employee", "Name", "Supervisors", "Stations", "Total", "W", "Y", "O", "R", "% Working"])
        for ro in data["ros"]:
            sm = ro["summary"]
            ws2.append([ro["employee_id"], ro["name"], ro["supervisor_count"], ro["station_count"],
                        sm["total"], sm["working"], sm["yellow"], sm["orange"], sm["red"], sm["pct_working"]])
        ws3 = wb.create_sheet("Departments")
        _hdr(ws3, ["RO", "Department", "Total", "W", "Y", "O", "R", "% Working"])
        for ro in data["ros"]:
            for r in ro["rings"]:
                ws3.append([ro["name"], r["name"], r["total"], r["working"], r["yellow"], r["orange"], r["red"], r["pct_working"]])

    # Flat assets sheet (drill to per-asset rows — F12)
    ws_assets = wb.create_sheet("Assets")
    _hdr(ws_assets, ["Asset #", "Asset Type", "Department", "Station", "Location", "Status", "Defective Since", "Hours Defective"])
    for a in asset_pool:
        t = U["type_by_id"].get(a.get("asset_type_id"), {})
        d = U["dept_by_id"].get(t.get("department_id"), {})
        s = U["station_by_id"].get(a.get("station_id"), {})
        l = U["location_by_id"].get(a.get("location_id"), {})
        ol_open = U["ol_by_asset"].get(str(a["_id"]))
        cls = _classify(a, ol_open)
        ds = ol_open.get("defective_since") if ol_open else a.get("defective_since")
        hours = ""
        if ds:
            if isinstance(ds, str):
                try:
                    ds_dt = datetime.fromisoformat(ds.replace("Z", "").replace("+00:00", ""))
                except Exception:
                    ds_dt = None
            else:
                ds_dt = ds.replace(tzinfo=None) if ds.tzinfo else ds
            if ds_dt:
                hours = round((now_ist() - ds_dt).total_seconds() / 3600, 1)
        ws_assets.append([
            a.get("asset_number"), t.get("name"), d.get("name"),
            s.get("name"), l.get("name"), cls.upper(),
            str(ds) if ds else "", hours,
        ])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"asset-health-report-{now_ist().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
