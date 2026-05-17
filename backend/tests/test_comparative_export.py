"""
Backend tests for Comparative Reports export (PDF + Excel).
Endpoints:
  POST /api/reports/comparative/export/pdf/{user_id}
  POST /api/reports/comparative/export/excel/{user_id}
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://asset-canvas.preview.emergentagent.com").rstrip("/")
SA_EMP = "SA001"
SUP_EMP = "SSE001"
PWD = "admin123"


# ── Auth helpers ─────────────────────────────────────────────────────────
def _login(emp_id, password=PWD):
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"employee_id": emp_id, "password": password}, timeout=30)
    assert r.status_code == 200, f"Login {emp_id} failed: {r.status_code} {r.text}"
    data = r.json()
    return data["token"], data["user"]


@pytest.fixture(scope="module")
def sa_ctx():
    token, user = _login(SA_EMP)
    return {"token": token, "user": user, "headers": {"Authorization": f"Bearer {token}"}}


@pytest.fixture(scope="module")
def sup_ctx():
    token, user = _login(SUP_EMP)
    return {"token": token, "user": user, "headers": {"Authorization": f"Bearer {token}"}}


@pytest.fixture(scope="module")
def departments(sa_ctx):
    r = requests.get(f"{BASE_URL}/api/departments", headers=sa_ctx["headers"], timeout=30)
    assert r.status_code == 200
    return r.json()


@pytest.fixture(scope="module")
def asset_types(sa_ctx):
    r = requests.get(f"{BASE_URL}/api/asset-types", headers=sa_ctx["headers"], timeout=30)
    assert r.status_code == 200
    return r.json()


def _default_body(**overrides):
    body = {
        "window_days": "90",
        "stat": "median",
        "dept_id": None,
        "asset_type_ids": None,
        "drill_state": {"level": "station", "parent_id": None, "parent_asset_type_id": None},
        "sections": {
            "card_a": True, "card_b": True, "card_c_current": True, "card_c_full": False,
            "defective": True, "remarks": True, "last_inspection": True,
        },
        "style": "detailed",
    }
    for k, v in overrides.items():
        if k == "sections":
            body["sections"].update(v)
        else:
            body[k] = v
    return body


# ── PDF export ───────────────────────────────────────────────────────────
class TestPdfExport:
    def test_pdf_default_export(self, sa_ctx):
        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/pdf/{uid}",
                          json=_default_body(), headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200, r.text
        assert "application/pdf" in r.headers.get("content-type", "")
        cd = r.headers.get("content-disposition", "")
        assert "filename=" in cd and ".pdf" in cd
        assert r.content[:8].startswith(b"%PDF-1."), f"Not a PDF: {r.content[:20]}"
        assert len(r.content) > 1000

    def test_pdf_compact_style(self, sa_ctx):
        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/pdf/{uid}",
                          json=_default_body(style="compact"), headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200
        assert r.content[:5] == b"%PDF-"

    def test_pdf_invalid_user(self, sa_ctx):
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/pdf/notanid",
                          json=_default_body(), headers=sa_ctx["headers"], timeout=30)
        assert r.status_code in (400, 404, 422)


# ── Excel export ─────────────────────────────────────────────────────────
def _open_xlsx(content):
    return load_workbook(io.BytesIO(content), read_only=False, data_only=False)


class TestExcelExport:
    def test_excel_default_all_sheets(self, sa_ctx):
        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
                          json=_default_body(), headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        cd = r.headers.get("content-disposition", "")
        assert "filename=" in cd and ".xlsx" in cd

        wb = _open_xlsx(r.content)
        sheets = set(wb.sheetnames)
        # default toggles (card_c_full=false): expect Drilldown not "Drilldown — Full"
        expected = {"Summary", "By Asset Type", "Peer Matrix", "Drilldown",
                    "Defective Only", "Last Inspections", "Remarks"}
        missing = expected - sheets
        assert not missing, f"Missing sheets: {missing}. Got: {sheets}"

    def test_excel_section_card_a_off(self, sa_ctx):
        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
                          json=_default_body(sections={"card_a": False}),
                          headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200
        wb = _open_xlsx(r.content)
        assert "By Asset Type" not in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_excel_section_defective_off(self, sa_ctx):
        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
                          json=_default_body(sections={"defective": False}),
                          headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200
        wb = _open_xlsx(r.content)
        assert "Defective Only" not in wb.sheetnames

    def test_excel_section_remarks_off(self, sa_ctx):
        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
                          json=_default_body(sections={"remarks": False}),
                          headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200
        wb = _open_xlsx(r.content)
        assert "Remarks" not in wb.sheetnames

    def test_excel_card_c_full(self, sa_ctx):
        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
                          json=_default_body(sections={"card_c_full": True, "card_c_current": False}),
                          headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200
        wb = _open_xlsx(r.content)
        assert "Drilldown — Full" in wb.sheetnames, f"Got: {wb.sheetnames}"
        ws = wb["Drilldown — Full"]
        # header row
        header = [c.value for c in ws[1]]
        assert "Asset Number" in header
        assert "Station" in header

    def test_excel_dept_filter_electrical(self, sa_ctx, departments, asset_types):
        elec = next((d for d in departments if "ELECT" in (d.get("name") or "").upper()), None)
        assert elec, "Electrical department not found"
        elec_id = elec["_id"]
        elec_type_names = {(t.get("name") or "").upper()
                           for t in asset_types if t.get("department_id") == elec_id}
        assert elec_type_names, "No electrical asset types"

        uid = sa_ctx["user"]["_id"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
                          json=_default_body(dept_id=elec_id), headers=sa_ctx["headers"], timeout=120)
        assert r.status_code == 200
        wb = _open_xlsx(r.content)
        ws = wb["By Asset Type"]
        # Read asset-type column (col B)
        seen = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                seen.append(str(row[1]).upper())
        # All seen types must be electrical
        for t in seen:
            assert t in elec_type_names, f"Non-electrical type in dept-filtered export: {t}"

    def test_excel_drill_state_location_summary(self, sa_ctx):
        # First, get a station id from comparative drill at station level
        uid = sa_ctx["user"]["_id"]
        r = requests.get(
            f"{BASE_URL}/api/reports/comparative/grouped/{uid}",
            params={"level": "station", "window_days": "90", "stat": "median"},
            headers=sa_ctx["headers"], timeout=60,
        )
        assert r.status_code == 200, r.text
        groups = r.json().get("groups") or []
        if not groups:
            pytest.skip("No station groups available")
        station_id = groups[0]["id"]
        station_label = groups[0]["label"]

        r = requests.post(
            f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
            json=_default_body(drill_state={"level": "location_summary",
                                            "parent_id": station_id,
                                            "parent_asset_type_id": None}),
            headers=sa_ctx["headers"], timeout=120,
        )
        assert r.status_code == 200
        wb = _open_xlsx(r.content)
        ws = wb["Drilldown"]
        # Row 1 = breadcrumb path; should reference station label
        path_cell = ws.cell(row=1, column=1).value or ""
        assert station_label in str(path_cell), f"Drill path cell does not include station: {path_cell}"


# ── SUP role anonymisation ───────────────────────────────────────────────
class TestSupAnonymisation:
    def test_excel_peer_matrix_anonymised(self, sup_ctx):
        uid = sup_ctx["user"]["_id"]
        own_name = sup_ctx["user"]["name"]
        r = requests.post(f"{BASE_URL}/api/reports/comparative/export/excel/{uid}",
                          json=_default_body(), headers=sup_ctx["headers"], timeout=120)
        assert r.status_code == 200, r.text
        wb = _open_xlsx(r.content)
        assert "Peer Matrix" in wb.sheetnames
        ws = wb["Peer Matrix"]
        # Collect supervisor labels (col A from row 2)
        labels = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and not str(row[0]).startswith("Note:"):
                labels.append(str(row[0]))
        # Self row should contain own name with ★
        self_rows = [l for l in labels if "★" in l]
        peer_rows = [l for l in labels if "★" not in l]
        if not self_rows and not peer_rows:
            pytest.skip("No peers/self in Peer Matrix")
        # Own row: must include real name
        if self_rows:
            assert any(own_name in r for r in self_rows), f"Self row missing own name: {self_rows}"
        # Peer rows: must NOT contain real names of peers (label 'Peer N')
        for pr in peer_rows:
            assert pr.lower().startswith("peer "), f"Peer row not anonymised: {pr}"
